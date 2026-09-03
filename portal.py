from flask import ( Flask, render_template,render_template_string, request, redirect, url_for, session, send_file, flash, jsonify)
import csv
import os
from flask_mail import Mail, Message
from datetime import datetime, timezone
from zoneinfo import ZoneInfo
from models import SystemNotification
from models import ClearanceRequest
from models import DownloadLog
from models import db, Term1, Term2, Term3
from io import BytesIO
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.pdfgen import canvas
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.lib.utils import ImageReader
import tempfile
from reportlab.lib.colors import HexColor
from flask_socketio import join_room
import math
import base64
from forms import UploadReceiptForm
from flask_login import LoginManager
from flask_migrate import Migrate 
from models import db, Student, PaymentHistory
from flask_login import LoginManager, UserMixin, login_user, login_required, current_user, logout_user
from werkzeug.utils import secure_filename
from flask_wtf import FlaskForm
from wtforms import FileField, SubmitField
from wtforms.validators import DataRequired
from decorators import admin_required
from models import Receipt
from models import User
from flask_mail import Mail, Message
import io
import matplotlib.pyplot as plt
from werkzeug.utils import secure_filename
from flask import send_file, make_response, send_from_directory
import mimetypes
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import ( SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer )
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph
from flask_socketio import SocketIO, emit
from flask_socketio import SocketIO, emit, join_room
from models import ReceiptNotification
import uuid
from PIL import Image
from models import DepartmentSignature
from sqlalchemy import String, func
from sqlalchemy import text
import shutil
import tempfile
from sqlalchemy import inspect, text
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.drawing.image import Image as ExcelImage, PILImage
from pdf2image import convert_from_path
from zoneinfo import ZoneInfo
import hashlib
import secrets
from datetime import datetime, timedelta
from models import Student, PasswordResetCode
from flask_mail import Message
import hmac
import time
from io import BytesIO
import qrcode
import random
from sqlalchemy import or_
from pdf2image import convert_from_bytes
from flask import (after_this_request, jsonify, redirect, send_file, session, url_for,)
import re



# Department passwords (change these to your desired secure passwords)
DEPARTMENT_PASSWORDS = {
    'reception': 'recept123',
    'library': 'lib456',
    'admission': 'adm789',
    'accounts': 'acc012',
    'systems': 'sys345',
    'adosa': 'ado678'
}
SECOND_DEPARTMENT_PASSWORDS = {
    'reception': 'RcpSecond!721',
    'library': 'LibSecond!834',
    'admission': 'AdmSecond!946',
    'accounts': 'AccSecond!258',
    'systems': 'SysSecond!369',
    'adosa': 'AdoSecond!475',
}
# Admin password for reports
ADMIN_PASSWORD = os.getenv(
    "ADMIN_PASSWORD",
    "admin123"
)  # Change this to a secure password
# ==========================================================
# APPLICATION CONFIGURATION
# ==========================================================

app = Flask(__name__)


# ==========================================================
# SECRET KEY
# ==========================================================

app.config["SECRET_KEY"] = os.getenv(
    "SECRET_KEY",
    "development-secret-key-change-this"
)


# ==========================================================
# DATABASE CONFIGURATION
# ==========================================================

database_url = os.getenv(
    "DATABASE_URL",
    "sqlite:///clearance_system.db"
)


# Render/PostgreSQL may provide postgres:// instead of
# postgresql://, so convert it automatically.

if database_url.startswith("postgres://"):

    database_url = database_url.replace(
        "postgres://",
        "postgresql://",
        1
    )


app.config["SQLALCHEMY_DATABASE_URI"] = database_url

app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
# ==========================================================
# INITIALIZE SHARED DATABASE
# ==========================================================

db.init_app(app)
socketio = SocketIO(app, cors_allowed_origins="*")
login_manager = LoginManager()
login_manager.init_app(app)


app.config.update(
    MAIL_SERVER="smtp.gmail.com",
    MAIL_PORT=587,
    MAIL_USE_TLS=True,
    MAIL_USE_SSL=False,
    MAIL_TIMEOUT=30,
    MAIL_USERNAME=os.getenv("MAIL_USERNAME"),
    MAIL_PASSWORD=os.getenv("MAIL_PASSWORD"),
    MAIL_DEFAULT_SENDER=os.getenv("MAIL_USERNAME"),
)


mail = Mail(app)

RECEIPT_ACTION_DENIED_MESSAGE = (
    "You are not allowed to perform this action only accounts "
    "and System can do so"
)


def hash_reset_code(code):
    return hashlib.sha256(code.encode("utf-8")).hexdigest()


def generate_reset_code():
    return f"{secrets.randbelow(1_000_000):06d}"


def clean_expired_reset_codes():
    PasswordResetCode.query.filter(
        PasswordResetCode.expires_at < datetime.utcnow()
    ).delete(synchronize_session=False)
    db.session.commit()

def send_email_with_retry(message, attempts=3):
    """Send an email using a fresh SMTP connection for each attempt."""
    last_error = None

    for attempt in range(1, attempts + 1):
        try:
            # A new connection is opened on every attempt. This avoids
            # reusing a connection that Gmail has already closed.
            with mail.connect() as connection:
                connection.send(message)

            app.logger.info(
                "Email sent successfully on attempt %s to %s",
                attempt,
                ", ".join(message.recipients or [])
            )
            return True

        except Exception as error:
            last_error = error
            app.logger.warning(
                "Email attempt %s/%s failed for %s: %s",
                attempt,
                attempts,
                ", ".join(message.recipients or []),
                error
            )

            if attempt < attempts:
                time.sleep(2)

    app.logger.error(
        "Email failed after %s attempts for %s: %s",
        attempts,
        ", ".join(message.recipients or []),
        last_error
    )
    return False



# The exact six departments allowed to sign a clearance.
DEPARTMENTS = {
    "accounts",
    "reception",
    "library",
    "admission",
    "systems",
    "adosa",
}


# ============================================================
# 2. EXISTING RECEIPT-CONFIRMATION EMAIL
# ============================================================
def send_confirmation_email(receipt):
    """Notify a student when Accounts confirms the uploaded receipt."""
    if not receipt:
        return False

    student = Student.query.filter_by(
        student_id=receipt.student_id
    ).first()

    if not student or not student.email:
        print("Receipt email not sent: student or email was not found.")
        return False

    student_name = getattr(
        receipt,
        "student_name",
        getattr(student, "name", student.email)
    )

    msg = Message(
        subject="Receipt Confirmed - Rockview University",
        recipients=[student.email],
    )

    msg.body = f"""
Hello {student_name},

Your receipt file:
{receipt.filename}

has been CONFIRMED by the Accounts Office.

Thank you.
Rockview University
"""

    try:
        mail.send(msg)
        print(f"Receipt confirmation email sent to {student.email}.")
        return True
    except Exception as error:
        print(f"Receipt confirmation email failed: {error}")
        return False


# ============================================================
# 3. DEPARTMENT-SIGNATURE EMAIL
# ============================================================
def send_department_signature_email(
    student,
    department,
    signed_count,
    total_departments=6,
    clearance_complete=False,
):
    """Notify the student immediately after a department signs."""
    if not student or not student.email:
        print("Signature email not sent: student email was not found.")
        return False

    department_name = department.title()
    student_name = getattr(
        student,
        "name",
        getattr(student, "student_name", student.email)
    )

    if clearance_complete:
        subject = "Clearance Complete - Rockview University"
        body_message = (
            "All six departments have signed your clearance. "
            "Your online university clearance process is now complete."
        )
    else:
        subject = (
            f"{department_name} Signed Your Clearance - "
            "Rockview University"
        )
        body_message = (
            f"The {department_name} department has successfully signed "
            "your clearance receipt."
        )

    msg = Message(
        subject=subject,
        recipients=[student.email],
    )

    msg.body = f"""
Dear {student_name},

{body_message}

Department progress: {signed_count} of {total_departments} departments completed.

Please log in to your Online University Clearance System dashboard to view the latest status.

Regards,
Online University Clearance System
Rockview University
"""

    try:
        mail.send(msg)
        print(
            f"{department_name} signature email sent to {student.email}."
        )
        return True
    except Exception as error:
        # The signature remains recorded even if Gmail is temporarily unavailable.
        print(
            f"{department_name} signature email failed for "
            f"{student.email}: {error}"
        )
        return False


# ============================================================
# 4. RECORD ONE DEPARTMENT SIGNATURE AND SEND EMAIL
# ============================================================
def record_department_signature(student_id, department):
    """Record a new departmental signature and notify the student."""
    department = department.lower().strip()

    if department not in DEPARTMENTS:
        raise ValueError(
            "Invalid department. Use Accounts, Reception, Library, "
            "Admission, Systems, or ADOSA."
        )

    student = Student.query.filter_by(
        student_id=str(student_id)
    ).first()

    if not student:
        raise ValueError("Student was not found.")

    existing_signature = DepartmentSignature.query.filter_by(
        student_id=str(student_id),
        department=department,
    ).first()

    if existing_signature:
        signed_count = DepartmentSignature.query.filter_by(
            student_id=str(student_id)
        ).count()

        return {
            "created": False,
            "complete": signed_count == len(DEPARTMENTS),
            "signed_count": signed_count,
            "total_departments": len(DEPARTMENTS),
            "email_sent": False,
        }

    signature = DepartmentSignature(
        student_id=str(student_id),
        department=department,
        signed_at=datetime.utcnow(),
    )
    db.session.add(signature)
    db.session.flush()

    signed_count = DepartmentSignature.query.filter_by(
        student_id=str(student_id)
    ).count()
    clearance_complete = signed_count == len(DEPARTMENTS)

    if clearance_complete:
        notification_title = "Clearance Complete"
        notification_message = (
            "All six departments have signed your clearance. "
            "Your clearance process is now complete."
        )
    else:
        notification_title = (
            f"{department.title()} Signed Your Clearance"
        )
        notification_message = (
            f"{department.title()} has signed your clearance. "
            f"Progress: {signed_count} of {len(DEPARTMENTS)} departments completed."
        )

    # Use the notification fields already present in your existing system.
    student_notification = SystemNotification(
        recipient_type="student",
        student_email=student.email,
        title=notification_title,
        message=notification_message,
        is_read=False,
        created_at=datetime.utcnow(),
    )
    db.session.add(student_notification)

    # Flush/record first, then send the immediate email.
    email_sent = send_department_signature_email(
        student=student,
        department=department,
        signed_count=signed_count,
        total_departments=len(DEPARTMENTS),
        clearance_complete=clearance_complete,
    )

    return {
        "created": True,
        "complete": clearance_complete,
        "signed_count": signed_count,
        "total_departments": len(DEPARTMENTS),
        "email_sent": email_sent,
    }


# ============================================================
# 5. ONE GENERIC SIGNATURE ROUTE
# ============================================================
# Use this route if your department dashboards can call one common endpoint.
# If you already have separate department-signing routes, call
# record_department_signature(student_id, department) inside those routes.

@app.route(
    "/department/<department>/sign/<int:student_id>",
    methods=["POST"],
)
def department_sign(department, student_id):
    try:
        result = record_department_signature(
            student_id=student_id,
            department=department,
        )

        db.session.commit()

        return jsonify({
            "success": True,
            "department": department,
            **result,
        })

    except ValueError as error:
        db.session.rollback()
        return jsonify({
            "success": False,
            "error": str(error),
        }), 400

    except Exception as error:
        db.session.rollback()
        return jsonify({
            "success": False,
            "error": "The signature could not be recorded.",
            "details": str(error),
        }), 500

def send_confirmation_email(receipt):

    if not receipt:
        return

    # Get student email from database instead of session (safer)
    student = Student.query.filter_by(student_id=receipt.student_id).first()

    if not student or not student.email:
        return

    msg = Message(
        subject="Receipt Confirmed",
        recipients=[student.email]
    )

    msg.body = f"""
Hello {receipt.student_name},

Your receipt file:
{receipt.filename}

has been CONFIRMED by the Accounts Office.

Thank you.
Rockview University
"""

    mail.send(msg)

migrate = Migrate(app, db)
app.jinja_env.globals.update(getattr=getattr)


def create_tables():
    db.create_all()


with app.app_context():
    create_tables()

ZAMBIA_TIMEZONE = ZoneInfo("Africa/Lusaka")


def format_zambia_time(value):
    if value is None:
        return ""

    if value.tzinfo is None:
        value = value.replace(tzinfo=timezone.utc)
    else:
        value = value.astimezone(timezone.utc)

    return value.astimezone(ZAMBIA_TIMEZONE).strftime(
        "%d/%m/%Y, %H:%M"
    )


def create_student(student_id, name, email, password):
    with app.app_context():
        existing = Student.query.get(student_id)
        if not existing:
            student = Student(id=student_id, name=name, email=email)
            student.set_password(password)
            db.session.add(student)
            db.session.commit()


def log_download(student, status):
    log = DownloadLog(
        student_id=student.id,
        student_name=student.name,
        program=student.program,
        intake=student.intake,
        status=status
    )
    db.session.add(log)
    db.session.commit()

# Utility function for amount_paid


def get_amount_paid_value(student):
    try:
        return int(student.amount_paid)
    except (TypeError, ValueError):
        return 0

# ============================
# Route: Sign Up
# ============================
@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

@app.route('/', methods=['GET', 'POST'])
def sign_up():
    if request.method == 'POST':
        student_id = request.form.get('student_id')
        full_name = request.form.get('full_name')
        email = request.form.get('email')
        password = request.form.get('password')
        program = request.form.get('program')
        intake = request.form.get('intake')

        if not all([student_id, full_name, email, password, program, intake]):
            flash('Please fill out all fields.')
            return redirect(url_for('sign_up'))

        existing_user = Student.query.filter(
            (Student.email == email) | (Student.id == student_id)
        ).first()

        if existing_user:
            flash('Email or Student ID already registered.')
            return redirect(url_for('sign_up'))

        new_student = Student(
            id=student_id,
            name=full_name,
            email=email,
            program=program,
            intake=intake
        )
        new_student.set_password(password)
        db.session.add(new_student)
        db.session.commit()

        flash('Registration successful! Please log in.')
        return redirect(url_for('login'))

    return render_template('sign_up.html')


@app.route('/print_students/<category>')
def print_students(category):
    def draw_header(c, page_width, page_height):
        # Draw the logo centered at the top
        try:
            logo_path = 'static/imgs/OIP.jpeg'  # Make sure this path iscorrect
            logo = ImageReader(logo_path)
            logo_width = 100
            logo_height = 50
            logo_x = (page_width - logo_width) / 2
            logo_y = page_height - logo_height - 20  # 20 units from top
            c.drawImage(logo, logo_x, logo_y, width=logo_width,
                        height=logo_height)
        except Exception:
            # If there's an error loading image, draw a rectangle placeholder
            c.setStrokeColorRGB(1, 0, 0)
            rect_x = (page_width - 100) / 2
            rect_y = page_height - 70
            c.rect(rect_x, rect_y, 100, 50)
            c.setFont("Times-Roman", 8)
            c.drawString(rect_x + 10, rect_y + 25, "Logo missing")
        # Draw institution name centered below the logo
        institution_name = "ROCKVIEW UNIVERSITY"
        c.setFont("Times-Bold", 16)
        text_width = c.stringWidth(institution_name, "Times-Bold", 16)
        c.drawString((page_width - text_width) / 2, page_height - 100,
                     institution_name)
        # Draw institution details centered below the name
        c.setFont("Times-Roman", 8)
        lines = [
            "Main Campus: 10-Miles, Lusaka",
            "Office Lines: 0211238065, Cell: +260 955 151 517, +260 967976961",
            "Vice Chancellor's line +260 973 184 162",
            "P.O Box 31108 Lusaka",
            "Web site: www.rockview.ac.zm",
            "Info@rockview.ac.zm",
            "Apply@rockview.ac.zm",
            "SCHOOLS OF EDUCATION, BUSINESS STUDIES AND AGRICULTURE SCIENCES",
            "NEW/RETURNING STUDENTS",
            "END OF TERM TEST/EXAM CLEARANCE FORM-UNDERGRADUATE STUDIES"
        ]
        y_start = page_height - 115  # start position below institution name
        for line in lines:
            line_width = c.stringWidth(line, "Times-Roman", 8)
            c.drawString((page_width - line_width) / 2, y_start, line)
            y_start -= 10  # spacing between lines

    # Simulate fetching students
    class Student:
        def __init__(self, id, name, email, amount_paid):
            self.id = id
            self.name = name
            self.email = email
            self.amount_paid = amount_paid

    students = [
        Student(1, "Alice Johnson", "alice@example.com", 25),
        Student(2, "Bob Smith", "bob@example.com", 50),
        Student(3, "Charlie Lee", "charlie@example.com", 100),
    ]

    # Determine category and title
    if category == '25':
        filtered_students = [s for s in students if s.amount_paid == 25]
        title = "Students who paid 25%"
    elif category == '50':
        filtered_students = [s for s in students if s.amount_paid == 50]
        title = "Students who paid 50%"
    elif category == '75':
        filtered_students = [s for s in students if s.amount_paid == 75]
        title = "Students who paid 75%"
    elif category == '100':
        filtered_students = [s for s in students if s.amount_paid == 100]
        title = "Students who paid 100%"
    else:
        return "Invalid category", 400

    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    # Draw the header (logo + institution info)
    draw_header(c, width, height)

    # Space for table (start lower to avoid overlapping header)
    y_position = height - 250  # adjust as needed for spacing

    # Title
    c.setFont("Times-Bold", 16)
    c.drawString(50, y_position, title)
    y_position -= 30

    # Table headers
    c.setFont("Times-Bold", 12)
    c.drawString(50, y_position, "Student ID")
    c.drawString(150, y_position, "Name")
    c.drawString(350, y_position, "Email")
    c.drawString(550, y_position, "Amount Paid")
    y_position -= 20

    # List students
    c.setFont("Times-Roman", 12)
    for s in filtered_students:
        if y_position < 50:
            c.showPage()
            draw_header(c, width, height)
            y_position = height - 250
            # Reprint headers after new page
            c.setFont("Times-Bold", 12)
            c.drawString(50, y_position, "Student ID")
            c.drawString(150, y_position, "Name")
            c.drawString(350, y_position, "Email")
            c.drawString(550, y_position, "Amount Paid")
            y_position -= 20
            c.setFont("Times-Roman", 12)
        c.drawString(50, y_position, str(s.id))
        c.drawString(150, y_position, s.name)
        c.drawString(350, y_position, s.email)
        c.drawString(550, y_position, f"{s.amount_paid}")
        y_position -= 20

    c.showPage()
    c.save()

    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name="students.pdf",
                     mimetype='application/pdf')


@app.route('/sign_with_password', methods=['POST'])
def sign_with_password():
    department = request.form.get('department')
    password = request.form.get('password')

    dept_map = {
        'reception': 'signed_reception',
        'library': 'signed_library',
        'admission': 'signed_admission',
        'accounts': 'signed_accounts',
        'systems': 'signed_systems',
        'adosa': 'signed_adosa'
    }

    if department not in dept_map:
        flash('Invalid department.')
        return redirect(url_for('dashboard'))

    expected_password = DEPARTMENT_PASSWORDS.get(department)
    if password != expected_password:
        flash('Incorrect department password.')
        return redirect(url_for('dashboard'))

    if department == 'accounts':
        student_id = session.get('student_id')
        if not student_id:
            return redirect(url_for('login'))
        student = Student.query.get(student_id)
        if not student:
            return redirect(url_for('login'))
        return redirect(url_for('accountant_pay', student_id=student.id))

    student_id = session.get('student_id')
    if not student_id:
        return redirect(url_for('login'))
    student = Student.query.get(student_id)
    if not student:
        return redirect(url_for('login'))

    setattr(student, dept_map[department], True)
    db.session.commit()

    flash(f'{department.capitalize()} signed successfully.')
    return redirect(url_for('dashboard'))


@app.route('/download_student_percentage')
def download_student_percentage():
    students = Student.query.all()
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []

    styles = getSampleStyleSheet()
    title = Paragraph("Students Payment Percentage Report", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 12))

    # Updated table header without 'Email'
    data = [['Student ID', 'Name', 'Program', 'Intake', 'Amount Paid (%)']]

    # Add student data rows, excluding email
    for s in students:
        data.append([
            s.id,
            s.name,
            s.program,
            s.intake,
            f"{s.amount_paid}%"  # Format percentage
        ])

    # Adjust colWidths: give wider space to 'Program' column
    col_widths = [70, 100, 200, 80, 80]  # Increased width for 'Program'

    # Create table with updated colWidths
    table = Table(data, colWidths=col_widths)
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
    ])
    table.setStyle(style)
    elements.append(table)

    # Build PDF
    doc.build(elements)
    buffer.seek(0)

    return send_file(buffer, as_attachment=True,
                     download_name='students_percentage_report.pdf',
                     mimetype='application/pdf')


@app.route('/download_payment_percentage/<int:percentage>')
def download_payment_percentage(percentage):
    # Fetch students with the specific amount paid percentage
    students = Student.query.filter_by(amount_paid=percentage).all()

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []

    styles = getSampleStyleSheet()
    title_text = f"Students Payment {percentage}% Report"
    title = Paragraph(title_text, styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 12))

    # Table header
    data = [['Student ID', 'Name', 'Program', 'Intake', 'Amount Paid (%)']]

    # Add student data rows
    for s in students:
        data.append([
            s.id,
            s.name,
            s.program,
            s.intake,
            f"{s.amount_paid}%"
        ])

    # Adjust colWidths for better program column width
    col_widths = [70, 100, 200, 80, 80]

    # Create table
    table = Table(data, colWidths=col_widths)
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
    ])
    table.setStyle(style)
    elements.append(table)

    # Build PDF
    doc.build(elements)
    buffer.seek(0)

    filename = f'students_payment_{percentage}_percent.pdf'
    return send_file(buffer, as_attachment=True, download_name=filename,
                     mimetype='application/pdf')


@app.route('/accounts_report')
def accounts_report():
    students = Student.query.order_by(Student.name.asc()).all()

    # PaymentHistory contains the real monetary amounts entered by Accounts.
    # The term tables are not included here because /add_payment saves the
    # same payment in both a term table and PaymentHistory.
    payment_total_rows = (
        db.session.query(
            PaymentHistory.student_id,
            func.coalesce(func.sum(PaymentHistory.amount), 0.0)
        )
        .group_by(PaymentHistory.student_id)
        .all()
    )

    # Map each student ID to the student's total monetary payments.
    student_payment_totals = {
        student_id: float(total or 0.0)
        for student_id, total in payment_total_rows
    }

    # Make sure students without a PaymentHistory record still appear as 0.
    for student in students:
        student_payment_totals.setdefault(student.id, 0.0)

    total_students = len(students)
    total_amount_paid = sum(student_payment_totals.values())

    # Preserve the previous filtered-list variable, but base it on actual
    # monetary totals rather than Student.amount_paid percentage.
    filtered_students = [
        student for student in students
        if 10000 <= student_payment_totals.get(student.id, 0.0) < 100000
    ]

    return render_template(
        'accounts_report.html',
        students=students,
        total_students=total_students,
        total_amount_paid=total_amount_paid,
        student_payment_totals=student_payment_totals,
        filtered_students=filtered_students,
    )


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        student_id = request.form.get('student_id')
        password = request.form.get('password')

        if not password:
            error = "Please enter your password."
            return render_template('login.html', error=error)

        student = Student.query.get(student_id)
        if student:
            if student.check_password(password):
                session['student_id'] = student.id
                return redirect(url_for('dashboard'))
        error = "Invalid Student ID or Password"
        return render_template('login.html', error=error)
    return render_template('login.html')

from flask import session, redirect, url_for, render_template, flash
from datetime import datetime

@app.route('/dashboard')
def dashboard():
    student_id = session.get('student_id')

    if not student_id:
        flash("Please login first")
        return redirect(url_for('login'))

    student = Student.query.get(student_id)

    if not student:
        flash("Student not found")
        return redirect(url_for('login'))

    # ===============================
    # Progress Calculation
    # ===============================
    progress = student.progress_percentage()

    # ===============================
    # Ensure signatures field exists
    # ===============================
    signatures = getattr(student, 'signatures', None)

    if signatures is None:
        signatures = {
            "Reception": "Pending",
            "Library": "Pending",
            "Admission": "Pending",
            "Accounts": "Pending",
            "Systems": "Pending",
            "ADOSA": "Pending"
        }
        student.signatures = signatures
        db.session.commit()

    # ===============================
    # Helper: check if all signed
    # ===============================
    def all_signed(sigs):
        return all(status == "Signed" for status in sigs.values())

    show_download = all_signed(signatures)

    # ===============================
    # 🔔 Receipt Notifications
    # ===============================
    notifications = Receipt.query.filter(
        Receipt.student_id == student.id,
        Receipt.status != 'pending'
    ).order_by(Receipt.timestamp.desc()).all()
    notifications = SystemNotification.query.filter(
    SystemNotification.recipient_id == str(student.id)
      ).order_by(
    SystemNotification.id.desc()
    ).all()
    # ===============================
    # FINAL RENDER
    # ===============================
    return render_template(
        'dashboard.html',
        student=student,
        progress=progress,
        signatures=signatures,
        show_download=show_download,
        notifications=notifications
    )

@app.route('/sign/<department>')
def sign_department(department):
    student_id = session.get('student_id')
    if not student_id:
        return redirect(url_for('login'))
    student = Student.query.get(student_id)
    if not student:
        return redirect(url_for('login'))

    department_map = {
        'reception': ('signed_reception', None),
        'library': ('signed_library', 'signed_reception'),
        'admission': ('signed_admission', 'signed_library'),
        'accounts': ('signed_accounts', 'signed_admission'),
        'systems': ('signed_systems', 'signed_accounts'),
        'adosa': ('signed_adosa', 'signed_systems')
    }

    if department.lower() not in department_map:
        return "Invalid department", 400

    attr_name, prev_attr = department_map[department.lower()]

    if prev_attr and not getattr(student, prev_attr):
        return '''
            <script>alert("You can't sign. Previous department
            has not signed."); window.history.back();</script>
        '''

    setattr(student, attr_name, True)
    db.session.commit()

    return redirect(url_for('dashboard'))

# SPECIAL ROUTE FOR ACCOUNTANT TO ENTER PAYMENT AMOUNT


@app.route(
    "/accountant_payment/<student_id>",
    methods=["GET", "POST"]
)
def accountant_payment(student_id):

    student = Student.query.get_or_404(student_id)

    if request.method == "POST":

        # Save payment percentage
        payment_option = request.form.get("payment_option")
        student.payment_percentage = payment_option

        db.session.commit()

        flash(
            "Payment percentage saved successfully. Please sign electronically.",
            "success"
        )

        # Redirect to the electronic signature page
        return redirect(
            url_for(
                "sign_student",
                student_id=student.id,
                department="accounts"
            )
        )

    # Detect the student's school/programme for the payment-period selector.
    # Humanities students receive Term 1–3; Health Sciences students receive
    # Semester 1–2 in accountant_pay.html.
    normalized_program = (
        str(student.program or "").strip().lower()
        .replace("&", "and")
        .replace(".", "")
        .replace("-", "_")
        .replace("/", "_")
        .replace(" ", "_")
    )

    is_health_sciences = any(
        keyword in normalized_program
        for keyword in (
            "health_science",
            "healthscience",
            "nurse",
            "nursing",
            "cog",
            "c_o_g",
            "clinical_officer",
        )
    )

    # GET REQUEST
    return render_template(
        "accountant_pay.html",
        student=student,
        is_health_sciences=is_health_sciences
    )    
# Helper function to generate clearance number
def generate_clearance_number(student_id):
    year = datetime.now().year
    # Assuming student_id is a string, include year and a random number for uniqueness
    random_no = random.randint(1000, 9999)
    return f"RVU-{year}-{student_id}-{random_no}"

# Helper function to generate verification code
def generate_verification_code():
    return ''.join(random.choices('ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789', k=8))



def flatten_pdf_to_image_pdf(vector_pdf_bytes, page_size):
    """
    Convert the already-generated clearance PDF into an image-only PDF.

    The visual appearance remains the same, but normal PDF text and table
    objects are removed. The final PDF contains only a high-resolution page
    image, including the QR code.
    """
    try:
        rendered_pages = convert_from_bytes(
            vector_pdf_bytes,
            dpi=220,
            fmt='png',
            thread_count=1,
        )
    except Exception as render_error:
        raise RuntimeError(
            'The PDF could not be flattened. Install Poppler and make sure '
            'pdf2image can find pdftoppm on this computer.'
        ) from render_error

    if not rendered_pages:
        raise RuntimeError('The clearance PDF produced no rendered pages.')

    image_pdf_buffer = BytesIO()
    image_pdf = canvas.Canvas(
        image_pdf_buffer,
        pagesize=page_size,
        pageCompression=1,
    )

    page_width, page_height = page_size

    for rendered_page in rendered_pages:
        page_image_buffer = BytesIO()
        rendered_page.save(
            page_image_buffer,
            format='PNG',
            optimize=True,
        )
        page_image_buffer.seek(0)

        image_pdf.drawImage(
            ImageReader(page_image_buffer),
            0,
            0,
            width=page_width,
            height=page_height,
            preserveAspectRatio=False,
            mask='auto',
        )
        image_pdf.showPage()

    image_pdf.save()
    image_pdf_buffer.seek(0)
    return image_pdf_buffer


# ================================================================
# REQUIRED IMPORTS
# ================================================================

# ================================================================
# FLATTEN THE GENERATED PDF INTO AN IMAGE-ONLY PDF
# ================================================================

def flatten_clearance_pdf(pdf_bytes, page_size):
    """Return a high-resolution image-only PDF."""

    poppler_path = os.environ.get("POPPLER_PATH", "").strip()

    convert_options = {
        "dpi": 220,
        "fmt": "png",
        "thread_count": 1,
    }

    # If POPPLER_PATH is empty, pdf2image searches the system PATH.
    if poppler_path:
        convert_options["poppler_path"] = poppler_path

    rendered_pages = convert_from_bytes(
        pdf_bytes,
        **convert_options,
    )

    if not rendered_pages:
        raise RuntimeError("No page was rendered from the clearance PDF.")

    image_pdf_buffer = BytesIO()
    image_pdf = canvas.Canvas(
        image_pdf_buffer,
        pagesize=page_size,
        pageCompression=1,
    )

    page_width, page_height = page_size

    for rendered_page in rendered_pages:
        page_image_buffer = BytesIO()
        rendered_page.save(
            page_image_buffer,
            format="PNG",
            optimize=True,
        )
        page_image_buffer.seek(0)

        image_pdf.drawImage(
            ImageReader(page_image_buffer),
            0,
            0,
            width=page_width,
            height=page_height,
            preserveAspectRatio=False,
            mask="auto",
        )
        image_pdf.showPage()

    image_pdf.save()
    image_pdf_buffer.seek(0)
    return image_pdf_buffer


# ================================================================
# COMPLETE DOWNLOAD ROUTE
# ================================================================



@app.route("/verify-clearance/<clearance_number>")
def verify_clearance(clearance_number):
    """Verify a clearance document using its clearance number and QR code."""
    import hmac

    supplied_code = (request.args.get("code") or "").strip()

    student = Student.query.filter_by(
        clearance_number=clearance_number
    ).first()

    verified = bool(
        student
        and supplied_code
        and student.verification_code
        and hmac.compare_digest(
            str(student.verification_code),
            supplied_code,
        )
    )

    return render_template(
        "verify_clearance.html",
        verified=verified,
        student=student if verified else None,
    )


@app.route("/download")
def download():
    student_id = session.get("student_id")

    if not student_id:
        return redirect(url_for("login"))

    student = db.session.get(Student, student_id)

    if not student:
        return redirect(url_for("login"))

    # ------------------------------------------------------------
    # ALL SIX DEPARTMENTS MUST SIGN BEFORE DOWNLOAD
    # ------------------------------------------------------------

    signatures = [
        student.signed_reception,
        student.signed_library,
        student.signed_admission,
        student.signed_accounts,
        student.signed_systems,
        student.signed_adosa,
    ]

    if not all(signatures):
        return jsonify({
            "status": "pending",
            "message": (
                "You can't download the clearance form until every "
                "department has signed."
            ),
        })

    # ------------------------------------------------------------
    # SECURITY INFORMATION
    # ------------------------------------------------------------

    if not student.clearance_number:
        student.clearance_number = generate_clearance_number(student.id)

    if not student.verification_code:
        student.verification_code = generate_verification_code()

    student.pdf_generated_date = datetime.now()
    db.session.commit()

    # ------------------------------------------------------------
    # CREATE THE ORIGINAL LANDSCAPE PDF
    # ------------------------------------------------------------

    page_size = landscape(letter)
    page_width, page_height = page_size
    buffer = BytesIO()

    pdf = canvas.Canvas(
        buffer,
        pagesize=page_size,
        pageCompression=1,
    )

    left_margin = 32
    right_margin = page_width - 32

    # ------------------------------------------------------------
    # LOCAL HELPERS
    # ------------------------------------------------------------

    def clean_text(value, fallback="Pending"):
        text = str(value or "").strip()
        return text if text else fallback

    def format_signed_time(value):
        if not value:
            return "Pending"

        if isinstance(value, datetime):
            return value.strftime("%d/%m/%Y %H:%M:%S")

        return str(value)

    def draw_existing_image(path_value, x, y, width, height):
        if not path_value:
            return

        relative_path = str(path_value).replace("/", os.sep)
        full_path = os.path.join(app.root_path, relative_path)

        if not os.path.exists(full_path):
            return

        try:
            pdf.drawImage(
                full_path,
                x,
                y,
                width=width,
                height=height,
                preserveAspectRatio=True,
                anchor="c",
                mask="auto",
            )
        except Exception as image_error:
            app.logger.warning(
                "Could not draw PDF image %s: %s",
                full_path,
                image_error,
            )

    def draw_cell_text(
        value,
        x,
        cell_width,
        baseline,
        font="Times-Roman",
        size=8,
        limit=22,
    ):
        text = clean_text(value)

        if len(text) > limit:
            text = text[:limit - 1] + "."

        pdf.setFont(font, size)
        pdf.drawCentredString(
            x + (cell_width / 2),
            baseline,
            text,
        )

    # ------------------------------------------------------------
    # BACKGROUND WATERMARK
    # ------------------------------------------------------------

    pdf.saveState()
    pdf.setFont("Helvetica", 11)
    pdf.setFillColor(HexColor("#dedede"))

    for watermark_y in range(35, int(page_height), 55):
        for watermark_x in range(-40, int(page_width), 170):
            pdf.drawString(
                watermark_x,
                watermark_y,
                "ROCKVIEW UNIVERSITY",
            )

    pdf.restoreState()

    # ------------------------------------------------------------
    # LOGO ABOVE INSTITUTION DETAILS
    # ------------------------------------------------------------

    logo_path = os.path.join(
        app.root_path,
        "static",
        "imgs",
        "OIP.jpeg",
    )

    logo_width = 54
    logo_height = 40
    logo_x = (page_width - logo_width) / 2
    logo_y = page_height - 55

    if os.path.exists(logo_path):
        try:
            pdf.drawImage(
                ImageReader(logo_path),
                logo_x,
                logo_y,
                width=logo_width,
                height=logo_height,
                preserveAspectRatio=True,
                anchor="c",
                mask="auto",
            )
        except Exception as logo_error:
            app.logger.warning(
                "Could not draw institution logo: %s",
                logo_error,
            )

    # ------------------------------------------------------------
    # INSTITUTION DETAILS BELOW THE LOGO
    # ------------------------------------------------------------

    y = logo_y - 17
    pdf.setFillColor(colors.black)
    pdf.setFont("Times-Bold", 15)
    pdf.drawCentredString(
        page_width / 2,
        y,
        "ROCKVIEW UNIVERSITY",
    )

    institution_lines = [
        "Main Campus: 10 Miles, Lusaka",
        "Office Lines: 0211238065 | +260955151517 | +260967976961",
        "Vice Chancellor: +260973184162",
        "P.O Box 31108 Lusaka",
        "Website: www.rockview.ac.zm",
        "Email: info@rockview.ac.zm",
    ]

    pdf.setFont("Times-Roman", 9)

    for institution_line in institution_lines:
        y -= 11
        pdf.drawCentredString(
            page_width / 2,
            y,
            institution_line,
        )

    # ------------------------------------------------------------
    # PROGRAM-SPECIFIC CLEARANCE HEADINGS
    # ------------------------------------------------------------

    # ------------------------------------------------------------
    # PROGRAM-SPECIFIC CLEARANCE HEADINGS
    # ------------------------------------------------------------
    # The request route stores the latest school and clearance selection in
    # the session. Read those values first so an old Humanities programme
    # value cannot override a current Health Sciences request.

    raw_program = (
        session.get("clearance_program")
        or getattr(student, "program", "")
        or ""
    )

    raw_clearance_type = (
        session.get("clearance_type")
        or getattr(student, "clearance_type", "")
        or ""
    )

    normalized_program = (
        str(raw_program).strip().lower()
        .replace("&", "and")
        .replace(".", "")
        .replace("-", "_")
        .replace("/", "_")
        .replace(" ", "_")
    )

    normalized_clearance_type = (
        str(raw_clearance_type).strip().lower()
        .replace("&", "and")
        .replace(".", "")
        .replace("-", "_")
        .replace("/", "_")
        .replace(" ", "_")
    )

    is_health_sciences = any(
        keyword in normalized_program
        for keyword in (
            "health_science",
            "healthscience",
            "nurse",
            "nursing",
            "cog",
            "c_o_g",
            "clinical_officer",
        )
    )

    health_sciences_title_map = {
        "mock": [
            "SCHOOL OF HEALTH SCIENCES",
            "MOCK EXAMINATION CLEARANCE FORM-UNDERGRADUATE STUDIES",
        ],
        "cat1": [
            "NEW / RETURNING STUDENTS",
            "SCHOOL OF HEALTH SCIENCES",
            "CAT1 / MOCK EXAMINATIONS",
            "STUDENTS CLEARANCE FORM-UNDERGRADUATE STUDIES",
        ],
        "cat2": [
            "NEW / RETURNING STUDENTS",
            "SCHOOL OF HEALTH SCIENCES",
            "CAT2 / EXAMINATION",
            "STUDENTS CLEARANCE FORM-UNDERGRADUATE STUDIES",
        ],
        "cat2_examination": [
            "NEW / RETURNING STUDENTS",
            "SCHOOL OF HEALTH SCIENCES",
            "CAT2 / EXAMINATION",
            "STUDENTS CLEARANCE FORM-UNDERGRADUATE STUDIES",
        ],
        "end_of_semester_examination": [
            "NEW / RETURNING STUDENTS",
            "SCHOOL OF HEALTH SCIENCES",
            "END OF SEMISTER EXAMINATION",
            "STUDENTS CLEARANCE FORM-UNDERGRADUATE STUDIES",
        ],
    }

    if is_health_sciences:
        heading_lines = health_sciences_title_map.get(
            normalized_clearance_type,
            [
                "SCHOOL OF HEALTH SCIENCES",
                "CLEARANCE FORM-UNDERGRADUATE STUDIES",
            ],
        )
    else:
        humanities_title_map = {
            "mid_term": [
                "NEW / RETURNING STUDENTS",
                "SCHOOL OF EDUCATION, HUMANITIES,",
                "AGRICULTURE SCIENCES AND BUSINESS STUDIES",
                "MID-TERM TESTS CLEARANCE FORM - UNDERGRADUATE STUDIES",
            ],
            "end_of_term": [
                "NEW / RETURNING STUDENTS",
                "SCHOOL OF EDUCATION, HUMANITIES,",
                "AGRICULTURE SCIENCES AND BUSINESS STUDIES",
                "END OF TERM TEST/ EXAM CLEARANCE FORM - UNDERGRADUATE STUDIES",
            ],
            "examination": [
                "NEW / RETURNING STUDENTS",
                "SCHOOL OF EDUCATION, HUMANITIES,",
                "AGRICULTURE SCIENCES AND BUSINESS STUDIES",
                "EXAMINATION CLEARANCE FORM - UNDERGRADUATE STUDIES",
            ],
        }

        heading_lines = humanities_title_map.get(
            normalized_clearance_type,
            [
                "NEW / RETURNING STUDENTS",
                "SCHOOL OF EDUCATION, HUMANITIES,",
                "AGRICULTURE SCIENCES AND BUSINESS STUDIES",
                "EXAMINATION CLEARANCE FORM - UNDERGRADUATE STUDIES",
            ],
        )

    y -= 17
    pdf.setFont("Times-Bold", 11)

    for heading_line in heading_lines:
        pdf.drawCentredString(
            page_width / 2,
            y,
            heading_line,
        )
        y -= 14

    y -= 2

    # ------------------------------------------------------------
    # STUDENT DETAILS
    # ------------------------------------------------------------

    y -= 22
    pdf.setFont("Times-Bold", 10)
    pdf.drawString(
        left_margin,
        y,
        f"Student Name: {clean_text(student.name, '')}",
    )

    y -= 13
    pdf.drawString(
        left_margin,
        y,
        f"Student ID: {clean_text(student.id, '')}",
    )

    y -= 13
    pdf.drawString(
        left_margin,
        y,
        f"Programme: {clean_text(student.program, '')}",
    )

    y -= 13
    pdf.drawString(
        left_margin,
        y,
        f"Intake: {clean_text(student.intake, '')}",
    )

    # The standalone horizontal line above the table was removed.
    # Keep spacing only; no pdf.line() is used here.
    y -= 25

    # ------------------------------------------------------------
    # PAYMENT DISPLAY
    # ------------------------------------------------------------

    raw_payment_value = getattr(
        student,
        "amount_paid",
        None,
    )

    raw_payment_status = getattr(
        student,
        "amount_paid_status",
        None,
    )

    payment_labels = {
        25: "Below 50%",
        50: "50%",
        75: "Above 50%",
        100: "100%",
    }

    payment_display = None

    if raw_payment_value is not None:
        try:
            numeric_payment_value = int(float(raw_payment_value))
            payment_display = payment_labels.get(
                numeric_payment_value,
                f"{numeric_payment_value}%",
            )
        except (TypeError, ValueError):
            payment_display = str(raw_payment_value).strip()

    if not payment_display and raw_payment_status:
        status_labels = {
            "50": "50%",
            "50%": "50%",
            "above_50": "Above 50%",
            "above 50%": "Above 50%",
            "below_50": "Below 50%",
            "below 50%": "Below 50%",
            "100": "100%",
            "100%": "100%",
        }

        status_text = str(raw_payment_status).strip().lower()
        payment_display = status_labels.get(
            status_text,
            str(raw_payment_status).strip(),
        )

    if not payment_display:
        payment_display = "Pending"

    # ------------------------------------------------------------
    # EIGHT-COLUMN TABLE
    # STEP | DEPARTMENT | REAM | T-SHIRT | COMMENT |
    # SIGNATURE | OFFICER | TIME
    # ------------------------------------------------------------

    table_x = left_margin
    table_top = y
    row_height = 32

    col_widths = [
        32,
        157,
        52,
        63,
        52,
        90,
        102,
        92,
    ]

    headers = [
        "STEP",
        "DEPARTMENT",
        "REAM",
        "T-SHIRT",
        "COMMENT",
        "SIGNATURE",
        "OFFICER",
        "TIME",
    ]

    department_rows = [
        (
            "1",
            "Reception",
            student.reception_clear,
            student.reception_clear,
            student.reception_clear,
            student.reception_signature,
            student.reception_officer_name,
            student.reception_signed_time,
        ),
        (
            "2",
            "Library",
            student.library_clear,
            student.library_clear,
            student.library_clear,
            student.library_signature,
            student.library_officer_name,
            student.library_signed_time,
        ),
        (
            "3",
            "Admission",
            student.admissions_clear,
            student.admissions_clear,
            student.admissions_clear,
            student.admissions_signature,
            student.admissions_officer_name,
            student.admissions_signed_time,
        ),
        (
            "4",
            "Accounts Office",
            student.accounts_clear,
            student.accounts_clear,
            payment_display,
            student.accounts_signature,
            student.accounts_officer_name,
            student.accounts_signed_time,
        ),
        (
            "5",
            "Systems",
            student.systems_clear,
            student.systems_clear,
            student.systems_clear,
            student.systems_signature,
            student.systems_officer_name,
            student.systems_signed_time,
        ),
        (
            "6",
            "Assistant Dean Student Affairs",
            student.adosa_clear,
            student.adosa_clear,
            student.adosa_clear,
            student.adosa_signature,
            student.adosa_officer_name,
            student.adosa_signed_time,
        ),
    ]

    total_width = sum(col_widths)
    total_rows = 1 + len(department_rows)
    table_height = total_rows * row_height
    table_bottom = table_top - table_height

    # Outer table border.
    pdf.setStrokeColor(colors.black)
    pdf.rect(
        table_x,
        table_bottom,
        total_width,
        table_height,
    )

    # Vertical lines.
    current_x = table_x

    for width in col_widths[:-1]:
        current_x += width
        pdf.line(
            current_x,
            table_top,
            current_x,
            table_bottom,
        )

    # Horizontal lines inside the table only.
    for row_number in range(1, total_rows):
        line_y = table_top - (row_number * row_height)
        pdf.line(
            table_x,
            line_y,
            table_x + total_width,
            line_y,
        )

    # Header background.
    pdf.setFillColor(HexColor("#d9e8f5"))
    pdf.rect(
        table_x,
        table_top - row_height,
        total_width,
        row_height,
        fill=1,
        stroke=0,
    )
    pdf.setFillColor(colors.black)

    # Header text.
    current_x = table_x
    pdf.setFont("Times-Bold", 8)

    for index, header in enumerate(headers):
        pdf.drawCentredString(
            current_x + (col_widths[index] / 2),
            table_top - 21,
            header,
        )
        current_x += col_widths[index]

    # Body rows.
    current_top = table_top - row_height

    for (
        step,
        department_name,
        ream_value,
        t_shirt_value,
        comment_value,
        signature_value,
        officer_name,
        signed_time,
    ) in department_rows:
        row_center = current_top - (row_height / 2) - 3
        x = table_x

        # STEP.
        draw_cell_text(
            step,
            x,
            col_widths[0],
            row_center,
            size=9,
            limit=5,
        )
        x += col_widths[0]

        # DEPARTMENT.
        department_text = clean_text(
            department_name,
            "",
        )

        if len(department_text) > 25:
            department_text = department_text[:24] + "."

        pdf.setFont("Times-Roman", 8)
        pdf.drawString(
            x + 3,
            row_center,
            department_text,
        )
        x += col_widths[1]

        # REAM.
        draw_existing_image(
            ream_value,
            x + 3,
            current_top - 28,
            col_widths[2] - 6,
            23,
        )
        x += col_widths[2]

        # T-SHIRT.
        # The Accounts row now receives the normal clearance mark here.
        draw_existing_image(
            t_shirt_value,
            x + 3,
            current_top - 28,
            col_widths[3] - 6,
            23,
        )
        x += col_widths[3]

        # COMMENT / ACCOUNTS PAYMENT STATUS.
        # The Accounts payment percentage is drawn as text in the next cell.
        if step == "4":
            draw_cell_text(
                comment_value,
                x,
                col_widths[4],
                row_center,
                font="Times-Bold",
                size=7,
                limit=14,
            )
        else:
            draw_existing_image(
                comment_value,
                x + 3,
                current_top - 28,
                col_widths[4] - 6,
                23,
            )
        x += col_widths[4]

        # SIGNATURE.
        draw_existing_image(
            signature_value,
            x + 3,
            current_top - 29,
            col_widths[5] - 6,
            25,
        )
        x += col_widths[5]

        # OFFICER.
        draw_cell_text(
            officer_name,
            x,
            col_widths[6],
            row_center,
            size=7,
            limit=17,
        )
        x += col_widths[6]

        # TIME.
        draw_cell_text(
            format_signed_time(signed_time),
            x,
            col_widths[7],
            row_center,
            size=6,
            limit=20,
        )

        current_top -= row_height

    # ------------------------------------------------------------
    # NOTE BELOW TABLE
    # ------------------------------------------------------------

    footer_y = table_bottom - 18
    pdf.setFillColor(colors.black)
    pdf.setFont("Times-Bold", 8)
    pdf.drawString(
        left_margin,
        footer_y,
        "PLEASE NOTE THAT ALL THE ABOVE OFFICES MUST SIGN.",
    )

    footer_y -= 14
    pdf.drawString(
        left_margin,
        footer_y,
        "DEAN - UNDERGRADUATE STUDIES",
    )

    # ------------------------------------------------------------
    # QR CODE — CENTERED BELOW THE CLEARANCE TABLE
    # ------------------------------------------------------------

    # Keep the QR code below the table instead of placing it at a fixed
    # right-hand position. This keeps it aligned even when heading text or
    # table content changes. The gap is approximately 1.5 line-spacings.
    # Keep the QR compact and immediately below the table so all four
    # corners remain visible above the footer when the PDF is scanned.
    qr_width = 64
    qr_height = 64
    qr_gap_below_table = 8
    qr_x = (page_width - qr_width) / 2
    qr_y = table_bottom - qr_height - qr_gap_below_table

    try:
        qr_path = generate_qr_code(student)

        if qr_path and os.path.exists(qr_path):
            pdf.drawImage(
                ImageReader(qr_path),
                qr_x,
                qr_y,
                width=qr_width,
                height=qr_height,
                preserveAspectRatio=True,
                anchor="c",
                mask="auto",
            )

            pdf.setFillColor(colors.black)
            pdf.setFont("Times-Bold", 8)
            pdf.drawCentredString(
                page_width / 2,
                qr_y - 12,
                "SCAN TO VERIFY",
            )
    except Exception as qr_error:
        app.logger.warning(
            "QR code error: %s",
            qr_error,
        )

    # ------------------------------------------------------------
    # FOOTER
    # ------------------------------------------------------------

    pdf.setFont("Times-Italic", 8)
    pdf.drawString(
        left_margin,
        21,
        "Generated by Rockview University Online Clearance System",
    )

    # ------------------------------------------------------------
    # FINISH THE VECTOR PDF
    # ------------------------------------------------------------

    pdf.showPage()
    pdf.save()
    buffer.seek(0)

    # ------------------------------------------------------------
    # CONVERT TO IMAGE-ONLY PDF
    # ------------------------------------------------------------

    try:
        secure_buffer = flatten_clearance_pdf(
            buffer.getvalue(),
            page_size,
        )
    except Exception as flatten_error:
        app.logger.exception(
            "Could not create image-only clearance PDF"
        )
        return jsonify({
            "success": False,
            "message": (
                "The clearance PDF could not be secured. "
                "Ensure Poppler is installed and available."
            ),
            "error": str(flatten_error),
        }), 500

    filename = (
        f"Clearance_Form_"
        f"{student.name}_"
        f"{student.id}.pdf"
    )

    # ------------------------------------------------------------
    # PREPARE THE DOWNLOAD RESPONSE BEFORE RESETTING THE RECORD
    # ------------------------------------------------------------
    # Keep the PDF in memory and register the reset as an after-request
    # callback. This ensures Flask prepares a real attachment response first.
    secure_buffer.seek(0)

    @after_this_request
    def reset_student_after_download(response):
        try:
            student.signed_reception = False
            student.signed_library = False
            student.signed_admission = False
            student.signed_accounts = False
            student.signed_systems = False
            student.signed_adosa = False
            db.session.commit()
        except Exception:
            db.session.rollback()
            app.logger.exception(
                "Could not reset student after clearance PDF response"
            )
        return response

    # ------------------------------------------------------------
    # RETURN THE PDF AS A VISIBLE BROWSER DOWNLOAD
    # ------------------------------------------------------------
    response = send_file(
        secure_buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="application/pdf",
        conditional=False,
        max_age=0,
        etag=False,
    )

    response.headers["Content-Disposition"] = (
        f'attachment; filename="{filename}"'
    )
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"

    return response



# ================================================================
# QR CODE GENERATOR
# ================================================================

def generate_qr_code(student):
    import qrcode

    # The first line is a clickable verification URL. The remaining lines
    # preserve the original readable student details inside the QR payload.
    verification_url = url_for(
        "verify_clearance",
        clearance_number=student.clearance_number,
        code=student.verification_code,
        _external=True,
    )

    generated_text = (
        student.pdf_generated_date.strftime("%d %B %Y %H:%M")
        if student.pdf_generated_date
        else "Not available"
    )

    qr_text = (
        f"{verification_url}\n\n"
        "ROCKVIEW UNIVERSITY\n"
        f"Student Name: {student.name}\n"
        f"Student ID: {student.id}\n"
        f"Programme: {student.program}\n"
        f"Intake: {student.intake}\n"
        f"Clearance Number: {student.clearance_number}\n"
        f"Verification Code: {student.verification_code}\n"
        f"Generated: {generated_text}"
    )

    qr = qrcode.QRCode(
        version=None,
        error_correction=qrcode.constants.ERROR_CORRECT_H,
        box_size=12,
        border=5,
    )

    qr.add_data(qr_text)
    qr.make(fit=True)

    qr_image = qr.make_image(
        fill_color="black",
        back_color="white",
    ).convert("RGB")

    temporary_file = tempfile.NamedTemporaryFile(
        suffix=".png",
        delete=False,
    )
    temporary_file.close()

    qr_image.save(
        temporary_file.name,
        format="PNG",
        optimize=False,
    )

    return temporary_file.name



@app.route('/record_success_download', methods=['POST'])
def record_success_download():
    student_id = session.get('student_id')
    if not student_id:
        return jsonify({'status': 'error', 'message': 'Not logged in'}), 401
    student = Student.query.get(student_id)
    if not student:
        return jsonify({'status': 'error', 'message': 'Student not found'}),
    404
    # Log success with current timestamp
    log = DownloadLog(
        student_id=student.id,
        student_name=student.name,
        program=student.program,
        intake=student.intake,
        status='success',
        timestamp=datetime.now()
    )
    try:
        db.session.add(log)
        db.session.commit()
        return jsonify({'status': 'success'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/get_success_logs', methods=['GET'])
def get_success_logs():
    logs = DownloadLog.query.filter_by(status='success').order_by
    (DownloadLog.timestamp.desc()).all()
    logs_data = [
        {
            'student_id': log.student_id,
            'student_name': log.student_name,
            'program': log.program,
            'intake': log.intake,
            'time_downloaded': log.timestamp.strftime('%Y-%m-%d %H:%M:%S')
        }
        for log in logs
    ]
    return jsonify(logs_data)


@app.route('/reset_signatures', methods=['POST'])
def reset_signatures():
    student_id = request.form.get('student_id')
    if not student_id:
        return '', 400
    student = Student.query.get(student_id)
    if not student:
        return '', 404
    # Reset all signature fields
    student.signed_reception = False
    student.signed_library = False
    student.signed_admission = False
    student.signed_accounts = False
    student.signed_systems = False
    student.signed_adosa = False
    db.session.commit()
    return '', 204

@app.route('/view_students')
def view_students():
    students = Student.query.all()
    return render_template('students.html', students=students)

# ============================
# Store logged-in department info in session
# ============================

@app.route('/department_login', methods=['GET', 'POST'])
def department_login():
    if request.method == 'POST':
        dept_id = (
            request.form.get('department_id') or ''
        ).strip().lower()

        password = request.form.get('password') or ''
        second_password = request.form.get('second_password') or ''
        officer_name = (
            request.form.get('officer_name') or ''
        ).strip()

        # Validate department ID.
        if dept_id not in DEPARTMENT_PASSWORDS:
            flash('Invalid Department ID', 'error')
            return redirect(url_for('department_login'))

        # Validate the original department password.
        expected_password = DEPARTMENT_PASSWORDS[dept_id]

        if password != expected_password:
            flash('Incorrect password', 'error')
            return redirect(url_for('department_login'))

        # Validate that the second password exists for this department.
        expected_second_password = SECOND_DEPARTMENT_PASSWORDS.get(dept_id)

        if expected_second_password is None:
            flash(
                'Second password is not configured for this department.',
                'error'
            )
            return redirect(url_for('department_login'))

        # Validate the department-specific second password.
        if second_password != expected_second_password:
            flash(
                'Incorrect second-layer security password',
                'error'
            )
            return redirect(url_for('department_login'))

        # Validate the officer name used by the signature and PDF system.
        if not officer_name:
            flash(
                'Please enter the officer full name before logging in.',
                'error'
            )
            return redirect(url_for('department_login'))

        # Save the exact values used by the signature route.
        session['department_id'] = dept_id
        session['officer_name'] = officer_name
        session['department_authenticated'] = True
        session.modified = True

        return redirect(url_for('department_dashboard'))

    return render_template('department_login.html')

@app.route('/department_logout')
def department_logout():
    session.pop('department_id', None)
    return redirect(url_for('department_login'))


@app.route('/department_dashboard')
def department_dashboard():
    dept_id = session.get('department_id')
    if not dept_id:
        return redirect(url_for('department_login'))

    search_query = request.args.get('search_query', '').strip()

    students = Student.query.all()

    students_list = []
    for s in students:
        students_list.append({
            'id': s.id,
            'name': s.name,
            'program': s.program,
            'intake': s.intake,
            'email': s.email,
            'signed_' + dept_id: getattr(s, 'signed_' + dept_id, False)
        })

    if search_query:
        students_list = [
            s for s in students_list
            if search_query.lower() in str(s['id']).lower() or search_query.lower() in s['email'].lower()
        ]

    # Pass dept_id to template to control button visibility
    return render_template(
        'department_dashboard.html',
        students=students_list,
        department_id=dept_id,
        request=request
    )
def save_signature_image(signature_data, department):

    signature_data = signature_data.split(",")[1]

    image_bytes = base64.b64decode(signature_data)

    folder = os.path.join(
        app.root_path,
        "static",
        "signatures",
        department
    )

    os.makedirs(folder, exist_ok=True)

    filename = f"{uuid.uuid4().hex}.png"

    filepath = os.path.join(folder, filename)

    with open(filepath, "wb") as f:
        f.write(image_bytes)

    return os.path.join(
        "static",
        "signatures",
        department,
        filename
    )
# ==========================================================
# SAVE HANDWRITTEN APPROVAL IMAGE
# ==========================================================
def save_approval_image(image_data, department):

    import os
    import uuid
    import base64

    # Remove the data:image/png;base64, prefix
    image_data = image_data.split(",")[1]

    image_bytes = base64.b64decode(image_data)

    folder = os.path.join(

        app.root_path,

        "static",

        "approval_marks"

    )

    os.makedirs(folder, exist_ok=True)

    filename = f"{department}_{uuid.uuid4().hex}.png"

    full_path = os.path.join(

        folder,

        filename

    )

    with open(full_path, "wb") as f:

        f.write(image_bytes)

    return os.path.join(

        "static",

        "approval_marks",

        filename

    )    

from flask import send_from_directory
import os

@app.route('/download_receipt/<filename>')
def download_receipt(filename):

    upload_folder = os.path.join(app.root_path, "uploads")

    return send_from_directory(
        upload_folder,
        filename,
        as_attachment=True
    )

@app.route('/receipt/<int:receipt_id>/delete', methods=['POST'])
def delete_receipt(receipt_id):
    # Only Accounts and System may delete receipts.
    if not accounts_or_system_can_manage_receipts():
        return deny_receipt_action()

    receipt = Receipt.query.get_or_404(receipt_id)

    # Your existing route allowed students to delete their own receipt.
    # That permission is intentionally removed here because this route
    # must be restricted to Accounts and System only.

    file_name = getattr(receipt, "filename", None)
    if not file_name:
        file_name = getattr(receipt, "file_name", None)

    upload_folder = os.path.join(app.root_path, "uploads")
    file_path = os.path.join(upload_folder, file_name) if file_name else None

    if file_path and os.path.exists(file_path):
        os.remove(file_path)

    db.session.delete(receipt)
    db.session.commit()

    return jsonify({
        "success": True,
        "message": "Receipt deleted successfully",
        "receipt_id": receipt.id,
    })


@app.route('/admin/receipts_dashboard')
def admin_receipts_dashboard():
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_login'))

    total = Receipt.query.count()
    pending = Receipt.query.filter_by(status='pending').count()
    approved = Receipt.query.filter_by(status='approved').count()
    rejected = Receipt.query.filter_by(status='rejected').count()

    return render_template(
        'admin_receipts_dashboard.html',
        total=total,
        pending=pending,
        approved=approved,
        rejected=rejected
    )


@app.route('/edit_student/<student_id>', methods=['GET', 'POST'])
def edit_student(student_id):
    # Fetch the student from the database
    student = Student.query.filter_by(id=student_id).first()
    if not student:
        flash("Student not found.", "error")
        return redirect(url_for('department_dashboard'))
    if request.method == 'POST':
        # Update student details from form data
        student.name = request.form.get('name')
        student.program = request.form.get('program')
        student.intake = request.form.get('intake')
        # Add other fields if any
        try:
            db.session.commit()
            flash("Student details updated successfully.", "success")
            return redirect(url_for('department_dashboard'))
        except Exception as e:
            db.session.rollback()
            flash("Error updating student: " + str(e), "error")
    # Render an edit form template
    return render_template('edit_student.html', student=student)


@app.route('/delete_student/<student_id>', methods=['POST'])
def delete_student(student_id):
    student = Student.query.filter_by(id=student_id).first()
    if not student:
        flash("Student not found.", "error")
        return redirect(url_for('department_dashboard'))
    try:
        db.session.delete(student)
        db.session.commit()
        flash("Student deleted successfully.", "success")
    except Exception as e:
        db.session.rollback()
        flash("Error deleting student: " + str(e), "error")
    return redirect(url_for('department_dashboard'))


@app.route('/record_download', methods=['POST'])
def record_download():
    student_id = session.get('student_id')
    if not student_id:
        return jsonify({'status': 'error', 'message': 'Not logged in'}), 401
    student = Student.query.get(student_id)
    if not student:
        return jsonify({'status': 'error', 'message': 'Student not found'}),
    404
    try:
        log = DownloadLog(
            student_id=student.id,
            student_name=student.name,
            program=student.program,
            intake=student.intake,
            status='success',
            timestamp=datetime.now()
        )
        db.session.add(log)
        db.session.commit()
        return jsonify({'status': 'success'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500
# ============================
# Run the app
# ============================

@app.route('/api/progress')
def api_progress():
    # Fetch department progress
    departments_progress = {
        'reception': 0,
        'library': 0,
        'admission': 0,
        'accounts': 0,
        'systems': 0,
        'adosa': 0
    }
    
    total_students = Student.query.count()

    if total_students == 0:
        total_students = 1  # prevent division by zero

    for dept in departments_progress.keys():
        signed_attr = 'signed_' + dept
        signed_count = Student.query.filter(getattr(Student, signed_attr) == True).count()
        departments_progress[dept] = (signed_count / total_students) * 100

    # Fetch individual students' progress
    students = Student.query.all()
    students_progress = [
        {
            'id': s.id,
            'name': s.name,
            'progress': (s.signed_reception + s.signed_library + s.signed_admission +
                         s.signed_accounts + s.signed_systems + s.signed_adosa) / 6 * 100
        } for s in students
    ]

    return jsonify({
        'departments': departments_progress,
        'students': students_progress
    })

@app.route('/reset_passwords')
def reset_passwords():
    with app.app_context():
        students = Student.query.all()
        for student in students:
            student.set_password('password123')  # Reset to default password
        db.session.commit()
        return "All student passwords have been reset to 'password123'.", 200


# Admin logout route
@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route('/pay', methods=['POST'])
def pay():
    student_id = session.get('student_id')
    payment_amount = float(request.form.get('amount'))

    student = Student.query.get(student_id)
    if student:
        student.amount_paid += payment_amount
        db.session.commit()
        # After updating, redirect or return updated progress
        return redirect(url_for('dashboard'))
    else:
        flash('Student not found')
        return redirect(url_for('dashboard'))

@app.route('/student_payment_progress/<student_id>')
def student_payment_progress(student_id):
    student = Student.query.get(student_id)
    if not student:
        return "Student not found", 404

    normalized_program = (str(student.program or '').strip().lower()
                          .replace('&', 'and').replace('.', '')
                          .replace('-', '_').replace('/', '_').replace(' ', '_'))
    is_health_sciences = any(key in normalized_program for key in (
        'health_science', 'healthscience', 'nurse', 'nursing', 'cog',
        'c_o_g', 'clinical_officer'
    ))

    # Keep the existing percentage chart for Humanities. For Health Sciences,
    # the chart summarizes the two stored semester amounts as a local finance
    # overview; no unsupported total-fee denominator is assumed.
    if is_health_sciences:
        semester1 = float(student.semester1_amount or 0)
        semester2 = float(student.semester2_amount or 0)
        values = [semester1, semester2]
        labels = ['Semester 1', 'Semester 2']
        colors_used = ['#198754', '#1261a0']
        chart_title = 'Health Sciences semester payments'
    else:
        progress = max(0, min(100, float(student.amount_paid or 0)))
        values = [progress, 100 - progress]
        labels = [f'Paid {progress:g}%', 'Remaining']
        colors_used = ['green', 'lightgrey']
        chart_title = 'Humanities term payment progress'

    fig, ax = plt.subplots(figsize=(4, 4))
    if sum(values) > 0:
        ax.pie(values, labels=labels, colors=colors_used, startangle=90, autopct='%1.1f%%')
    else:
        ax.pie([1], labels=['No payment recorded'], colors=['lightgrey'], startangle=90)
    ax.set_title(chart_title, fontsize=10)
    ax.axis('equal')

    buf = io.BytesIO()
    plt.savefig(buf, format='png', bbox_inches='tight')
    plt.close(fig)
    buf.seek(0)
    image_base64 = base64.b64encode(buf.read()).decode('utf-8')

    return render_template('payment_progress.html', image_base64=image_base64,
                           student=student, is_health_sciences=is_health_sciences)

@app.route('/payment-history/<student_id>')
def payment_history(student_id):
    student = Student.query.get_or_404(student_id)

    term1_payments = Term1.query.filter_by(
        student_id=student.id
    ).order_by(Term1.payment_date.desc()).all()

    term2_payments = Term2.query.filter_by(
        student_id=student.id
    ).order_by(Term2.payment_date.desc()).all()

    term3_payments = Term3.query.filter_by(
        student_id=student.id
    ).order_by(Term3.payment_date.desc()).all()

    return render_template(
        'payment_history.html',
        student=student,
        term1_payments=term1_payments,
        term2_payments=term2_payments,
        term3_payments=term3_payments,
    )




@app.route('/add_payment/<student_id>', methods=['POST'])
def add_payment(student_id):
    student = Student.query.get_or_404(student_id)

    # The Accounts form uses name="term" for both Humanities terms and
    # Health Sciences semesters. The fallback also accepts name="semester".
    selected_period = (
        request.form.get('term')
        or request.form.get('semester')
        or ''
    ).strip()
    amount_text = (request.form.get('amount') or '').strip()
    date_str = (request.form.get('payment_date') or '').strip()
    time_str = (request.form.get('payment_time') or '').strip()
    description = (request.form.get('description') or '').strip()

    allowed_periods = {
        'Term 1',
        'Term 2',
        'Term 3',
        'Semester 1',
        'Semester 2',
    }

    if selected_period not in allowed_periods:
        flash('Please select a valid term or semester.', 'danger')
        return redirect(url_for('accountant_payment', student_id=student.id))

    try:
        amount = float(amount_text)
        if amount < 0:
            raise ValueError
    except (TypeError, ValueError):
        flash('Please enter a valid non-negative payment amount.', 'danger')
        return redirect(url_for('accountant_payment', student_id=student.id))

    try:
        if time_str:
            payment_date = datetime.strptime(
                f'{date_str} {time_str}',
                '%Y-%m-%d %H:%M'
            )
        else:
            payment_date = datetime.strptime(date_str, '%Y-%m-%d')
    except ValueError:
        flash('Invalid payment date or time.', 'danger')
        return redirect(url_for('accountant_payment', student_id=student.id))

    # Preserve the existing Humanities period tables.
    period_table_record = None
    if selected_period == 'Term 1':
        period_table_record = Term1(
            student_id=student.id,
            amount=amount,
            payment_date=payment_date,
            description=description,
            status='Paid',
            payment_method='Manual'
        )
    elif selected_period == 'Term 2':
        period_table_record = Term2(
            student_id=student.id,
            amount=amount,
            payment_date=payment_date,
            description=description,
            status='Paid',
            payment_method='Manual'
        )
    elif selected_period == 'Term 3':
        period_table_record = Term3(
            student_id=student.id,
            amount=amount,
            payment_date=payment_date,
            description=description,
            status='Paid',
            payment_method='Manual'
        )

    if period_table_record is not None:
        db.session.add(period_table_record)

    # Health Sciences semester summaries are cumulative. Each payment is also
    # written to PaymentHistory so the semester table can show every payment.
    if selected_period == 'Semester 1':
        student.semester1_amount = (student.semester1_amount or 0) + amount
        student.semester1_payment_date = payment_date
        student.semester1_payment_status = 'Paid'
    elif selected_period == 'Semester 2':
        student.semester2_amount = (student.semester2_amount or 0) + amount
        student.semester2_payment_date = payment_date
        student.semester2_payment_status = 'Paid'

    payment_record = PaymentHistory(
        student_id=student.id,
        amount=amount,
        payment_date=payment_date,
        description=description,
        payment_method='Manual',
        term=selected_period,
        status='Paid'
    )
    db.session.add(payment_record)
    db.session.commit()

    flash(
        f'{selected_period} payment of K {amount:,.2f} saved successfully.',
        'success'
    )
    return redirect(url_for('accountant_payment', student_id=student.id))



@app.route('/add_payment_option/<student_id>', methods=['POST'])
def add_payment_option(student_id):
    student = Student.query.get_or_404(student_id)

    payment_option = (
        request.form.get('payment_option') or ''
    ).strip().lower()

    payment_values = {
        '50': {
            'amount': 50,
            'label': '50%',
        },
        'above_50': {
            'amount': 75,
            'label': 'Above 50%',
        },
        'below_50': {
            'amount': 25,
            'label': 'Below 50%',
        },
        '100': {
            'amount': 100,
            'label': '100%',
        },
    }

    selected_payment = payment_values.get(payment_option)

    if selected_payment is None:
        flash(
            'Please select 50%, Above 50%, Below 50%, or 100%.',
            'danger'
        )
        return redirect(
            url_for(
                'accountant_payment',
                student_id=student.id
            )
        )

    # Numeric value retained for existing database/PDF compatibility.
    student.amount_paid = selected_payment['amount']

    # Readable wording used by the PDF.
    if hasattr(student, 'amount_paid_status'):
        student.amount_paid_status = selected_payment['label']

    db.session.commit()

    flash(
        'Payment recorded successfully. Please sign electronically.',
        'success'
    )

    return redirect(
        url_for(
            'sign_student',
            student_id=student.id,
            department='accounts'
        )
    )

    

    
@app.route('/receipts/<int:receipt_id>/view', methods=['POST'])
@login_required
@admin_required
def update_receipt_status(receipt_id):
    receipt = Receipt.query.get_or_404(receipt_id)
    receipt.status = 'viewed'
    db.session.commit()
    return redirect(url_for('view_receipts'))

@app.route('/receipts/<int:receipt_id>/feedback', methods=['POST'])
@login_required
@admin_required
def send_feedback(receipt_id):
    receipt = Receipt.query.get_or_404(receipt_id)
    # Send feedback to student via email or in-app notification
    return redirect(url_for('view_receipts'))
import os
from werkzeug.utils import secure_filename
from flask import send_file, make_response, send_from_directory
import mimetypes

# Allowed formats (ADDED PDF SUPPORT)
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'pdf'}

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


@app.route('/upload_receipt', methods=['GET', 'POST'])
def upload_receipt():

    if request.method == 'POST':

        uploaded_file = request.files.get('receipt')

        if not uploaded_file or uploaded_file.filename == '':
            flash('No file uploaded')
            return redirect(url_for('upload_receipt'))

        if not allowed_file(uploaded_file.filename):
            flash('Only PDF, JPG, JPEG, PNG files are allowed')
            return redirect(url_for('upload_receipt'))

        # =========================
        # GET LOGGED-IN STUDENT
        # =========================
        student_id = session.get('student_id')

        if not student_id:
            return redirect(url_for('login'))

        student = Student.query.get(student_id)

        if not student:
            flash("Student not found")
            return redirect(url_for('login'))

        student_name = student.name
        student_email = student.email

        # =========================
        # AUTO RENAME FILE
        # =========================
        extension = uploaded_file.filename.rsplit('.', 1)[1].lower()
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        new_filename = f"{student_id}_{timestamp}.{extension}"

        # =========================
        # CREATE UPLOAD FOLDER
        # =========================
        upload_folder = os.path.join(app.root_path, 'uploads')

        if not os.path.exists(upload_folder):
            os.makedirs(upload_folder)

        upload_path = os.path.join(upload_folder, new_filename)
        uploaded_file.save(upload_path)

        # =========================
        # SAVE RECEIPT TO DB
        # =========================
        receipt = Receipt(
            student_id=student_id,
            student_name=student_name,
            filename=new_filename,
            status='Pending',
            timestamp=datetime.now()
        )

        db.session.add(receipt)
        db.session.commit()

        # =========================
        # SAVE SYSTEM NOTIFICATION
        # =========================
        system_notification = SystemNotification(
            recipient_type='accounts',
            recipient_id='accounts',
            title='New Receipt Uploaded',
            message=f"{student_name} ({student_id}) uploaded a new receipt",
            student_email=student.email
        )

        db.session.add(system_notification)

        # =========================
        # SAVE RECEIPT NOTIFICATION
        # =========================
        receipt_notification = ReceiptNotification(
            receipt_id=receipt.id,
            student_id=student.id,
            student_name=student.name,
            student_email=student.email,
            message=f"{student.name} ({student.student_id}) uploaded a new receipt"
        )
        db.session.add(receipt_notification)
        db.session.commit()

        # =========================
        # REAL-TIME SOCKET ALERT
        # =========================
        socketio.emit(
            "new_receipt",
            {
                "student_name": student.name,
                "student_id": student.student_id if hasattr(student, 'student_id') else student.id,
                "student_email": student.email,
                "message": f"{student.name} uploaded a receipt",
                "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            },
            room="accounts"
        )

        flash('Receipt uploaded successfully')
        return redirect(url_for('view_receipts'))

    return render_template('upload_form.html')



@app.route('/view_receipts')
def view_receipts():

    student_id = session.get('student_id')

    if not student_id:
        return redirect(url_for('login'))

    receipts = Receipt.query.filter(
        Receipt.student_id == int(student_id)
    ).order_by(
        Receipt.id.desc()
    ).all()

    print("SESSION STUDENT ID =", student_id)
    print("RECEIPTS FOUND =", len(receipts))

    return render_template(
        'view_receipts.html',
        receipts=receipts
    )
    # ---------------- SERVE RECEIPTS ----------------
from flask import send_from_directory
import os

@app.route('/receipt/<filename>')
def serve_receipt(filename):

    upload_folder = os.path.join(app.root_path, 'uploads')

    receipt = Receipt.query.filter_by(filename=filename).first()

    if session.get('admin_logged_in') and receipt:
        if receipt.status == "Pending":
            receipt.status = "Confirmed"
            db.session.commit()

            send_confirmation_email(receipt)  # 🔥 EMAIL SENT HERE

    return send_from_directory(upload_folder, filename)



@app.route('/admin/receipt/<int:receipt_id>/update', methods=['POST'])
def admin_update_receipt(receipt_id):
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_login'))

    receipt = Receipt.query.get_or_404(receipt_id)

    status = request.form.get('status')
    comment = request.form.get('admin_comment')

    receipt.status = status
    receipt.admin_comment = comment
    if status.lower() == "approved":

        notification = SystemNotification(
        recipient_type='student',
        recipient_id=str(receipt.student_id),
        title='Receipt Approved',
        message='Your receipt has been approved by Accounts.'
    )

    db.session.add(notification)

    db.session.commit()

    flash("Receipt updated successfully")
    return redirect(url_for('view_receipts'))


@app.route("/studentdashboard")
def studentdashboard():
    # ------------------------------
    # 1. Check login
    # ------------------------------
    if "student_id" not in session:
        return redirect(url_for("login"))

    student_id = session["student_id"]

    # ------------------------------
    # 2. Get student from DB
    # ------------------------------
    student = Student.query.get(student_id)
    if not student:
        flash("Student not found")
        return redirect(url_for("login"))

    
    # ------------------------------
    # 5. Ensure signatures exist
    # ------------------------------
    signatures = getattr(student, 'signatures', None)
    if signatures is None:
        signatures = {
            "Reception": "Pending",
            "Library": "Pending",
            "Admission": "Pending",
            "Accounts": "Pending",
            "Systems": "Pending",
            "ADOSA": "Pending"
        }
        student.signatures = signatures
        db.session.commit()

    # ------------------------------
    # 6. Helper: check if all signed
    # ------------------------------
    def all_signed(sigs):
        return all(status == "Signed" for status in sigs.values())

    show_download = all_signed(signatures)

    # ------------------------------
    # 7. Notifications (optional)
    # ------------------------------
    notifications = Receipt.query.filter(
        Receipt.student_id == student.id,
        Receipt.status != 'pending'
    ).order_by(Receipt.timestamp.desc()).all()

    # ------------------------------
    # 8. Render template
    # ------------------------------
    return render_template(
        "dashboard.html",
        student=student,
        signatures=signatures,
        show_download=show_download,
        notifications=notifications
    )

from flask_mail import Message
def send_confirmation_email(receipt):

    if not receipt:
        return

    student_email = session.get('student_email')

    if not student_email:
        return

    msg = Message(
        subject="Receipt Confirmed",
        recipients=[student_email]
    )

    msg.body = f"""
    Hello {receipt.student_name},

    Your receipt ({receipt.filename}) has been confirmed.

    Thank you.
    """

    mail.send(msg)

def notification_student_details(notification):
    """Read the student name and ID from the existing notification message."""
    message = notification.message or ""
    match = re.match(r'^(.+?)\s*\(([^)]+)\)', message)

    if match:
        return match.group(1).strip(), match.group(2).strip()

    return "Student", ""


@app.route("/api/accounts/notifications", methods=["GET"])
def accounts_notifications_api():
    notifications = (
        SystemNotification.query
        .filter_by(recipient_type="accounts")
        .order_by(
            SystemNotification.created_at.desc(),
            SystemNotification.id.desc()
        )
        .limit(200)
        .all()
    )

    result = []

    for notification in notifications:
        result.append({
            "id": notification.id,
            "student_email": notification.student_email or "",
            "title": notification.title or "New Clearance Request",
            "message": notification.message or "",
            "is_read": bool(notification.is_read),
            "created_at_zambia": format_zambia_time(
                notification.created_at
            )
        })

    return jsonify({
        "notifications": result
    })



@app.route('/accounts/notifications/<int:notification_id>/read', methods=['POST'])
def mark_notification_read(notification_id):
    """Mark one Accounts notification as read."""
    notification = SystemNotification.query.filter_by(
        id=notification_id,
        recipient_type='accounts'
    ).first_or_404()

    notification.is_read = True
    db.session.commit()

    return jsonify({
        'success': True,
        'id': notification_id
    })


@app.route('/request_clearance', methods=['GET', 'POST'])
def request_clearance():
    """Create or update a student's clearance request.

    Humanities clearance types:
        mid-term, end-of-term, examination

    Health Sciences clearance types:
        mock, cat1, cat2, end_of_semester_examination

    The selected clearance type is stored on Student.clearance_type.
    ClearanceRequest does not receive clearance_type as a constructor
    argument because that field is not defined on the ClearanceRequest model.
    """

    # ==========================================================
    # CHECK STUDENT LOGIN
    # ==========================================================
    student_id = session.get('student_id')

    if not student_id:
        flash('Please login first.', 'error')
        return redirect(url_for('login'))

    # ==========================================================
    # LOAD STUDENT
    # ==========================================================
    student = Student.query.get_or_404(student_id)

    # ==========================================================
    # HANDLE CLEARANCE REQUEST SUBMISSION
    # ==========================================================
    if request.method == 'POST':
        clearance_program = (
            request.form.get('clearance_program') or ''
        ).strip().lower()

        clearance_type = (
            request.form.get('clearance_type') or ''
        ).strip().lower()

        # Accept the programme values submitted by the HTML page.
        allowed_programmes = {
            'humanities',
            'health_sciences',
        }

        # Humanities options.
        humanities_types = {
            'mid-term',
            'end-of-term',
            'examination',
        }

        # Health Sciences options.
        health_sciences_types = {
            'mock',
            'cat1',
            'cat2',
            'end_of_semester_examination',
        }

        allowed_types = humanities_types | health_sciences_types

        # ======================================================
        # VALIDATE PROGRAMME
        # ======================================================
        if clearance_program not in allowed_programmes:
            flash(
                'Please select either School of Humanities or '
                'School of Health Sciences.',
                'error'
            )
            return redirect(url_for('request_clearance'))

        # ======================================================
        # VALIDATE CLEARANCE TYPE
        # ======================================================
        if clearance_type not in allowed_types:
            flash(
                'Please select a valid clearance request type.',
                'error'
            )
            return redirect(url_for('request_clearance'))

        # ======================================================
        # PREVENT CROSS-PROGRAMME SELECTIONS
        # ======================================================
        if (
            clearance_program == 'humanities'
            and clearance_type not in humanities_types
        ):
            flash(
                'The selected clearance type does not belong to '
                'School of Humanities.',
                'error'
            )
            return redirect(url_for('request_clearance'))

        if (
            clearance_program == 'health_sciences'
            and clearance_type not in health_sciences_types
        ):
            flash(
                'The selected clearance type does not belong to '
                'School of Health Sciences.',
                'error'
            )
            return redirect(url_for('request_clearance'))

        # ======================================================
        # SAVE THE SELECTED VALUES
        # ======================================================
        # Student is the authoritative model field for clearance_type.
        student.clearance_type = clearance_type

        # The PDF route reads these session values when choosing the title.
        session['clearance_program'] = clearance_program
        session['clearance_type'] = clearance_type
        session.modified = True

        # ======================================================
        # FIND OR CREATE CLEARANCE REQUEST
        # ======================================================
        existing_request = ClearanceRequest.query.filter_by(
            student_id=student_id
        ).first()

        if existing_request:
            existing_request.status = 'Pending'

            # Do not assign existing_request.clearance_type here because
            # ClearanceRequest has no clearance_type model column.
            if hasattr(existing_request, 'clearance_program'):
                existing_request.clearance_program = clearance_program

            if hasattr(existing_request, 'updated_at'):
                existing_request.updated_at = datetime.utcnow()

            # Remove old unread notifications for this student's previous
            # request before creating the fresh notifications below.
            SystemNotification.query.filter_by(
                student_email=student.email,
                title='New Clearance Request',
                is_read=False,
            ).delete(synchronize_session=False)

            request_record = existing_request

        else:
            # Only pass fields accepted by ClearanceRequest.
            # clearance_type belongs to Student, not ClearanceRequest.
            request_record = ClearanceRequest(
                student_id=student_id,
                status='Pending',
            )

            if hasattr(request_record, 'student_name'):
                request_record.student_name = student.name

            if hasattr(request_record, 'clearance_program'):
                request_record.clearance_program = clearance_program

            if hasattr(request_record, 'created_at'):
                request_record.created_at = datetime.utcnow()

            db.session.add(request_record)

        # ======================================================
        # CREATE FRESH DEPARTMENT NOTIFICATIONS
        # ======================================================
        department_names = [
            'reception',
            'library',
            'admission',
            'accounts',
            'systems',
            'adosa',
        ]

        programme_label = (
            'School of Health Sciences'
            if clearance_program == 'health_sciences'
            else 'School of Humanities'
        )

        clearance_labels = {
            'mid-term': 'Mid-Term Test',
            'end-of-term': 'End-of-Term Test',
            'examination': 'Examination',
            'mock': 'Mock Examination',
            'cat1': 'CAT1 / Mock Examinations',
            'cat2': 'CAT2 / Examination',
            'end_of_semester_examination': (
                'End-of-Semester Examination'
            ),
        }

        clearance_label = clearance_labels.get(
            clearance_type,
            clearance_type.replace('-', ' ').title(),
        )

        for department in department_names:
            notification = SystemNotification(
                recipient_type=department,
                recipient_id=department,
                title='New Clearance Request',
                message=(
                    f'{student.name} ({student.id}) requested clearance '
                    f'for {clearance_label} under {programme_label}.'
                ),
                student_email=student.email,
                is_read=False,
                created_at=datetime.utcnow(),
            )

            db.session.add(notification)

        # ======================================================
        # SAVE ONCE AND RETURN TO THE STUDENT DASHBOARD
        # ======================================================
        try:
            db.session.commit()
        except Exception:
            db.session.rollback()
            app.logger.exception(
                'Could not save clearance request for student %s',
                student.id,
            )
            flash(
                'The clearance request could not be saved. Please try again.',
                'error'
            )
            return redirect(url_for('request_clearance'))

        flash(
            f'{clearance_label} clearance request submitted successfully. '
            'Your request has been sent to all departments.',
            'success'
        )

        return redirect(url_for('dashboard'))

    # ==========================================================
    # DISPLAY CLEARANCE REQUEST PAGE
    # ==========================================================
    return render_template(
        'request_clearance.html',
        student=student,
    )
    
@app.route('/accounts_notifications')
def accounts_notifications():
    # The page is now populated by JavaScript through the JSON endpoint.
    return render_template('accounts_notifications.html')


@app.route('/finalize_accounts/<student_id>', methods=['POST'])
def finalize_accounts(student_id):

    student = Student.query.get_or_404(student_id)

    payment_option = request.form.get('payment_option')

    if not payment_option:

        flash(
            'Select payment percentage.',
            'danger'
        )

        return redirect(
            url_for(
                'accountant_payment',
                student_id=student.id
            )
        )

    # ==========================================================
    # SAVE PAYMENT PERCENTAGE
    # ==========================================================
    student.payment_percentage = payment_option

    # ==========================================================
    # DO NOT MARK AS SIGNED YET
    # THE OFFICER MUST FIRST DRAW THE E-SIGNATURE
    # ==========================================================

    notification = SystemNotification(

        recipient_id=str(student.id),

        title='Accounts Approval',

        message=(
            f'Accounts Department processed your payment. '
            f'Fees Paid: {payment_option}%. '
            f'Awaiting electronic signature.'
        )

    )

    db.session.add(notification)

    db.session.commit()

    flash(
        'Payment recorded successfully. Please provide your electronic signature.',
        'success'
    )

    # ==========================================================
    # REDIRECT TO THE ELECTRONIC SIGNATURE PAGE
    # ==========================================================
    return redirect(
        url_for(
            'sign_student',
            student_id=student.id,
            department='accounts'
        )
    )


def _receipt_file_candidates(filename):
    """Return the possible storage locations for an uploaded receipt."""
    if not filename:
        return []

    upload_folder = app.config.get("UPLOAD_FOLDER", "uploads")

    return [
        os.path.join(upload_folder, filename),
        os.path.join(app.root_path, upload_folder, filename),
        os.path.join(app.root_path, "uploads", filename),
        os.path.join(app.root_path, "static", "uploads", filename),
        os.path.join(app.root_path, "receipts", filename),
    ]


def _find_receipt_file(filename):
    """Find the physical receipt file on disk, if it exists."""
    for candidate in _receipt_file_candidates(filename):
        if os.path.isfile(candidate):
            return candidate

    return None


def _receipt_student_data(receipt):
    """Return safe display data for an active or trashed receipt."""
    student = Student.query.filter_by(
        id=str(receipt.student_id)
    ).first()

    return {
        "receipt_obj": receipt,
        "student_id": (
            str(student.id)
            if student
            else str(receipt.student_id or "")
        ),
        "name": (
            student.name
            if student and student.name
            else (receipt.student_name or "Student name unavailable")
        ),
        "email": student.email if student else "",
        "intake": student.intake if student else "",
        "program": student.program if student else "",
        "date_uploaded": (
            receipt.timestamp.strftime("%Y-%m-%d %H:%M:%S")
            if receipt.timestamp else ""
        ),
        "deleted_at": (
            receipt.deleted_at.strftime("%Y-%m-%d %H:%M:%S")
            if getattr(receipt, "deleted_at", None) else ""
        ),
        "deleted_by": getattr(receipt, "deleted_by", "") or "",
    }

@app.route("/view_all_receipts")
def view_all_receipts():
    search = request.args.get("search", "").strip().lower()
    requested_status = request.args.get("status", "").strip().lower()

    # Active table: deleted receipts are excluded.
    all_receipts = (
        Receipt.query
        .filter(Receipt.is_deleted.is_(False))
        .order_by(Receipt.id.desc())
        .all()
    )

    # Trash sidebar: deleted receipts remain available for review.
    deleted_receipts = (
        Receipt.query
        .filter(Receipt.is_deleted.is_(True))
        .order_by(Receipt.deleted_at.desc(), Receipt.id.desc())
        .all()
    )

    total_receipts = len(all_receipts)
    pending = sum(
        1 for receipt in all_receipts
        if str(receipt.status or "").strip().lower() == "pending"
    )
    approved = sum(
        1 for receipt in all_receipts
        if str(receipt.status or "").strip().lower() == "approved"
    )
    rejected = sum(
        1 for receipt in all_receipts
        if str(receipt.status or "").strip().lower() == "rejected"
    )

    receipts_data = []
    for receipt in all_receipts:
        receipt_status = str(receipt.status or "").strip().lower()
        if requested_status and receipt_status != requested_status:
            continue

        item = _receipt_student_data(receipt)
        if search:
            searchable = " ".join([
                item["student_id"], item["name"], item["email"],
                item["intake"], item["program"], receipt_status,
            ]).lower()
            if search not in searchable:
                continue
        receipts_data.append(item)

    trash_data = [_receipt_student_data(receipt) for receipt in deleted_receipts]

    return render_template(
        "view_all_receipts.html",
        receipts=receipts_data,
        trash_receipts=trash_data,
        trash_count=len(trash_data),
        total_receipts=total_receipts,
        pending=pending,
        approved=approved,
        rejected=rejected,
    )

@app.route('/accounts/receipt/<int:receipt_id>/trash', methods=['POST'])
def move_receipt_to_trash(receipt_id):
    if not accounts_or_system_can_manage_receipts():
        return deny_receipt_action()

    receipt = Receipt.query.get_or_404(receipt_id)

    if receipt.is_deleted:
        return jsonify({
            "success": True,
            "message": "Receipt is already in Trash",
            "receipt_id": receipt.id,
        })

    receipt.is_deleted = True
    receipt.deleted_at = datetime.utcnow()
    receipt.deleted_by = str(
        session.get("username")
        or session.get("user_id")
        or session.get("department_id")
        or "Accounts"
    )

    db.session.commit()

    return jsonify({
        "success": True,
        "message": "Receipt moved to Trash",
        "receipt_id": receipt.id,
    })


# ============================================================
# 6. RESTORE ONE RECEIPT
# ============================================================
@app.route("/receipt/<int:receipt_id>/restore", methods=["POST"])
def restore_receipt(receipt_id):
    receipt = Receipt.query.get_or_404(receipt_id)
    receipt.is_deleted = False
    receipt.deleted_at = None
    receipt.deleted_by = None
    db.session.commit()

    return jsonify({
        "success": True,
        "message": "Receipt restored",
        "receipt_id": receipt.id,
    })


# ============================================================
# 7. RESTORE ALL TRASHED RECEIPTS
# ============================================================
@app.route("/receipts/trash/restore-all", methods=["POST"])
def restore_all_receipts():
    deleted_receipts = Receipt.query.filter(
        Receipt.is_deleted.is_(True)
    ).all()

    for receipt in deleted_receipts:
        receipt.is_deleted = False
        receipt.deleted_at = None
        receipt.deleted_by = None

    db.session.commit()
    return jsonify({
        "success": True,
        "restored_count": len(deleted_receipts),
    })


# ============================================================
# 8. PERMANENT DELETE: ONLY ALLOW RECORDS ALREADY IN TRASH
# ============================================================
@app.route("/receipt/<int:receipt_id>/permanent-delete", methods=["POST"])
def permanently_delete_receipt(receipt_id):
    receipt = Receipt.query.get_or_404(receipt_id)

    if not receipt.is_deleted:
        return jsonify({
            "success": False,
            "message": "Only receipts already in Trash can be permanently deleted.",
        }), 400

    file_path = _find_receipt_file(receipt.filename)
    db.session.delete(receipt)
    db.session.commit()

    if file_path:
        try:
            os.remove(file_path)
        except OSError:
            app.logger.warning("Could not remove receipt file: %s", file_path)

    return jsonify({
        "success": True,
        "message": "Receipt permanently deleted",
        "receipt_id": receipt_id,
    })


# ============================================================
# 9. EXCEL BACKUP WITH ALL ACTIVE AND TRASHED RECEIPTS
# ============================================================
@app.route("/accounts/receipts/backup.xlsx")
def backup_all_receipts_excel():
    receipts = Receipt.query.order_by(Receipt.id.asc()).all()

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "All Receipt Records"

    headers = [
        "Receipt ID", "Record Location", "Student ID", "Student Name",
        "Student Email", "Program", "Intake", "Receipt Status",
        "Receipt Filename", "Date Uploaded", "Admin Comment",
        "Deleted At", "Deleted By", "Receipt Image / File",
    ]
    worksheet.append(headers)

    header_fill = PatternFill("solid", fgColor="003366")
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    temporary_pdf_dir = tempfile.mkdtemp(prefix="receipt_backup_")

    for receipt in receipts:
        item = _receipt_student_data(receipt)
        location = "Trash" if receipt.is_deleted else "Active"
        image_or_file = ""

        row = [
            receipt.id,
            location,
            item["student_id"],
            item["name"],
            item["email"],
            item["program"],
            item["intake"],
            receipt.status or "",
            receipt.filename or "",
            item["date_uploaded"],
            receipt.admin_comment or "",
            item["deleted_at"],
            item["deleted_by"],
            image_or_file,
        ]
        worksheet.append(row)
        current_row = worksheet.max_row

        file_path = _find_receipt_file(receipt.filename)
        if file_path and PILImage and receipt.filename:
            extension = os.path.splitext(receipt.filename)[1].lower()
            if extension in {".jpg", ".jpeg", ".png", ".gif", ".bmp"}:
                try:
                    excel_image = ExcelImage(file_path)
                    excel_image.width = 120
                    excel_image.height = 120
                    worksheet.add_image(excel_image, f"N{current_row}")
                    worksheet.row_dimensions[current_row].height = 95
                    worksheet.cell(current_row, 14).value = "Embedded image"
                except Exception as exc:
                    worksheet.cell(current_row, 14).value = f"Image unavailable: {exc}"
            elif extension == ".pdf":
                try:
                    pages = convert_from_path(
                        file_path,
                        first_page=1,
                        last_page=1,
                        dpi=120,
                    )
                    if pages:
                        preview_path = os.path.join(
                            temporary_pdf_dir,
                            f"receipt_{receipt.id}.png",
                        )
                        pages[0].save(preview_path, "PNG")
                        excel_image = ExcelImage(preview_path)
                        excel_image.width = 120
                        excel_image.height = 120
                        worksheet.add_image(excel_image, f"N{current_row}")
                        worksheet.row_dimensions[current_row].height = 95
                        worksheet.cell(current_row, 14).value = "Embedded PDF first page"
                    else:
                        worksheet.cell(current_row, 14).value = f"PDF: {receipt.filename}"
                except Exception as exc:
                    worksheet.cell(current_row, 14).value = f"PDF preview unavailable: {exc}"
            else:
                worksheet.cell(current_row, 14).value = f"File: {receipt.filename}"
        elif receipt.filename:
            worksheet.cell(current_row, 14).value = f"File not found: {receipt.filename}"

    widths = {
        "A": 12, "B": 16, "C": 16, "D": 25, "E": 30,
        "F": 24, "G": 18, "H": 16, "I": 30, "J": 22,
        "K": 35, "L": 22, "M": 18, "N": 22,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    # Temporary PDF preview images are no longer needed after workbook.save().
    # Remove only the temporary conversion directory created by this export.
    if "temporary_pdf_dir" in locals() and temporary_pdf_dir:
        shutil.rmtree(temporary_pdf_dir, ignore_errors=True)

    return send_file(
        output,
        as_attachment=True,
        download_name=(
            f"university_clearance_receipts_backup_"
            f"{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        ),
        mimetype=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
    )
# ==========================================================
# BACKUP ACTIVE RECEIPTS TO EXCEL
# ==========================================================

@app.route('/accounts/receipts/backup.xlsx')
def backup_receipts_to_excel():

    department = str(
        session.get('department_id')
        or session.get('department')
        or session.get('role')
        or ''
    ).strip().lower()

    # Only authorized departments can create backups
    if department not in {'accounts', 'system', 'systems'}:

        return jsonify({
            'success': False,
            'message': (
                'You are not allowed to download the receipts backup.'
            )
        }), 403


    # ======================================================
    # GET ACTIVE RECEIPTS ONLY
    # ======================================================

    active_receipts = Receipt.query.filter(
        Receipt.is_deleted.isnot(True)
    ).all()


    # ======================================================
    # CREATE EXCEL WORKBOOK
    # ======================================================

    workbook = Workbook()

    worksheet = workbook.active

    worksheet.title = "Active Receipts Backup"


    # ======================================================
    # EXCEL TITLE
    # ======================================================

    worksheet.append([
        "ROCKVIEW UNIVERSITY ONLINE CLEARANCE SYSTEM"
    ])

    worksheet.append([
        "ACTIVE RECEIPTS BACKUP"
    ])

    worksheet.append([])


    # ======================================================
    # TABLE HEADERS
    # ======================================================

    headers = [

        "Receipt ID",

        "Student ID",

        "Student Name",

        "Date Uploaded",

        "Status",

        "Comment",

        "Receipt Filename"

    ]


    worksheet.append(headers)


    # ======================================================
    # ADD RECEIPT DATA
    # ======================================================

    for receipt in active_receipts:

        worksheet.append([

            receipt.id,

            getattr(
                receipt,
                'student_id',
                ''
            ),

            getattr(
                receipt,
                'student_name',
                ''
            ),

            str(
                getattr(
                    receipt,
                    'uploaded_at',
                    ''
                )
                or
                getattr(
                    receipt,
                    'date_uploaded',
                    ''
                )
            ),

            getattr(
                receipt,
                'status',
                ''
            ),

            getattr(
                receipt,
                'comment',
                ''
            ),

            getattr(
                receipt,
                'filename',
                ''
            )

        ])


    # ======================================================
    # MAKE COLUMNS READABLE
    # ======================================================

    worksheet.column_dimensions["A"].width = 15

    worksheet.column_dimensions["B"].width = 18

    worksheet.column_dimensions["C"].width = 30

    worksheet.column_dimensions["D"].width = 25

    worksheet.column_dimensions["E"].width = 18

    worksheet.column_dimensions["F"].width = 40

    worksheet.column_dimensions["G"].width = 40


    # ======================================================
    # CREATE EXCEL FILE IN MEMORY
    # ======================================================

    excel_file = BytesIO()

    workbook.save(excel_file)

    excel_file.seek(0)


    # ======================================================
    # DOWNLOAD EXCEL FILE
    # ======================================================

    return send_file(

        excel_file,

        as_attachment=True,

        download_name=(
            "Rockview_University_Active_Receipts_Backup.xlsx"
        ),

        mimetype=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )

    )

from flask import send_from_directory
import os
from flask import send_from_directory
@app.route('/receipt/<filename>')
def view_receipt(filename):
    upload_folder = os.path.join(app.root_path, 'uploads')
    return send_from_directory(upload_folder, filename)

@app.route('/receipt/<int:receipt_id>/approve', methods=['POST'])
def approve_receipt(receipt_id):
    department = str(
        session.get('department_id')
        or session.get('department')
        or session.get('role')
        or ''
    ).strip().lower()

    if department not in {'accounts', 'system', 'systems'}:
        return jsonify({
            'success': False,
            'message': (
                'You are not allowed to perform this action only accounts '
                'and System can do so'
            )
        }), 403

    receipt = Receipt.query.get_or_404(receipt_id)

    # Update the receipt status without deleting the receipt record.
    receipt.status = 'Approved'
    receipt.admin_comment = 'Approved by Accounts/System'

    # Find only the latest unread Accounts notification
    # belonging to this specific student.
    notification = SystemNotification.query.filter(
        SystemNotification.recipient_type == 'accounts',
        SystemNotification.is_read.is_(False),
        SystemNotification.message.like(
            f'%({receipt.student_id})%'
        )
    ).order_by(
        SystemNotification.id.desc()
    ).first()

    # Reduce unread notifications by exactly one.
    if notification:
        notification.is_read = True

    # Commit both the Approved status and the notification read status.
    db.session.commit()

    return jsonify({
        'success': True,
        'status': 'Approved',
        'receipt_id': receipt.id,
        'notification_marked_read': notification is not None
    })


@app.route('/receipt/<int:receipt_id>/reject', methods=['POST'])
def reject_receipt(receipt_id):
    # Only Accounts and System may reject receipts.
    if not accounts_or_system_can_manage_receipts():
        return deny_receipt_action()

    receipt = Receipt.query.get_or_404(receipt_id)

    data = request.get_json(silent=True) or {}
    reason = str(data.get('reason', '')).strip()

    if not reason:
        return jsonify({
            "success": False,
            "message": "Please provide a rejection reason.",
        }), 400

    receipt.status = "Rejected"
    receipt.admin_comment = reason

    db.session.commit()

    return jsonify({
        "success": True,
        "status": "Rejected",
        "receipt_id": receipt.id,
    })



@app.route('/receipt/<int:receipt_id>/history')
def receipt_history(receipt_id):
    receipt = Receipt.query.get_or_404(receipt_id)
    history = {
        "id": receipt.id,
        "student_id": receipt.student_id,
        "student_name": receipt.student_name,
        "status": receipt.status,
        "comment": receipt.admin_comment,
        "uploaded": receipt.timestamp.strftime("%Y-%m-%d %H:%M:%S")
    }
    return jsonify(history)

from flask import jsonify

@socketio.on('join')
def handle_join(data):
    room = data.get('room')
    if room:
        join_room(room)

import re


@app.route('/receipt/notifications', methods=['GET'])
def receipt_notifications():
    # Get active receipts only. Trashed receipts are excluded.
    active_receipts = Receipt.query.filter(
        Receipt.is_deleted.is_(False)
    ).order_by(
        Receipt.timestamp.desc(),
        Receipt.id.desc()
    ).all()

    # Create a lookup of active receipts by student ID.
    receipts_by_student = {}
    for receipt in active_receipts:
        student_key = str(receipt.student_id or '').strip()
        if student_key:
            receipts_by_student.setdefault(student_key, []).append(receipt)

    # Get unread Accounts notifications only.
    unread_notifications = SystemNotification.query.filter(
        SystemNotification.recipient_type == 'accounts',
        SystemNotification.is_read.is_(False)
    ).order_by(
        SystemNotification.created_at.desc(),
        SystemNotification.id.desc()
    ).all()

    valid_notifications = []
    used_receipt_ids = set()

    for notification in unread_notifications:
        message = notification.message or ''

        # Your messages are in this format:
        # Student Name (StudentID) uploaded a receipt
        match = re.search(r'\(([^)]+)\)', message)
        if not match:
            continue

        student_key = match.group(1).strip()
        matching_receipts = receipts_by_student.get(student_key, [])

        # Match one unread notification to one existing receipt only.
        matched_receipt = None
        for receipt in matching_receipts:
            if receipt.id not in used_receipt_ids:
                matched_receipt = receipt
                break

        # Ignore duplicate/stale notifications that have no receipt match.
        if matched_receipt is None:
            continue

        used_receipt_ids.add(matched_receipt.id)
        valid_notifications.append(notification)

    messages = [
        notification.message or 'New receipt uploaded'
        for notification in valid_notifications
    ]

    latest_info = (
        valid_notifications[0].message
        if valid_notifications
        else 'No unread receipt notifications.'
    )

    return jsonify({
        # This is the unread notification count only.
        # It can never be greater than the number of active receipts.
        'new_count': len(valid_notifications),
        'messages': messages,
        'latest_info': latest_info,
    })

@app.route("/receipt/notifications/mark_read", methods=["POST"])
def mark_all_notifications_read():
    # Bulk update: mark all unread notifications as read
    ReceiptNotification.query.filter_by(is_read=False).update({ReceiptNotification.is_read: True})
    db.session.commit()
    return jsonify({"success": True, "message": "All notifications marked as read."})

@app.route("/accounts/unread_count")
def unread_count():

    count = ReceiptNotification.query.filter_by(
        is_read=False
    ).count()

    return jsonify({"count":count})

@app.route('/clear_accounts_notifications', methods=['POST'])
def clear_accounts_notifications():

    deleted = SystemNotification.query.filter(
        SystemNotification.recipient_type == 'accounts'
    ).delete(synchronize_session=False)

    db.session.commit()

    flash(f'{deleted} notification records deleted successfully.')

    return redirect(url_for('accounts_notifications'))
@app.route('/backup_accounts_notifications')
def backup_accounts_notifications():

    notifications = SystemNotification.query.filter_by(
        recipient_type='accounts'
    ).order_by(
        SystemNotification.id.desc()
    ).all()

    backup_folder = os.path.join(
        app.root_path,
        'backups'
    )

    if not os.path.exists(backup_folder):
        os.makedirs(backup_folder)

    filename = f"Accounts_Backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

    filepath = os.path.join(
        backup_folder,
        filename
    )

    with open(filepath, mode='w', newline='', encoding='utf-8') as file:

        writer = csv.writer(file)

        writer.writerow([
            "Title",
            "Message",
            "Status",
            "Date",
            "Student Email"
        ])

        for n in notifications:

            writer.writerow([
                n.title,
                n.message,
                "Approved" if n.is_read else "Pending",
                n.created_at.strftime('%d-%m-%Y %H:%M'),
                n.student_email
            ])

    return send_file(
        filepath,
        as_attachment=True
    )

@app.route("/receipts/trash/permanent-delete-all", methods=["POST"])
def permanently_delete_all_receipts():
    deleted_receipts = Receipt.query.filter(
        Receipt.is_deleted.is_(True)
    ).all()

    file_paths = []

    for receipt in deleted_receipts:
        file_path = _find_receipt_file(receipt.filename)
        if file_path:
            file_paths.append(file_path)

    deleted_count = len(deleted_receipts)

    try:
        for receipt in deleted_receipts:
            db.session.delete(receipt)

        db.session.commit()

    except Exception as error:
        db.session.rollback()
        app.logger.exception("Could not permanently delete all trashed receipts")
        return jsonify({
            "success": False,
            "message": "The trashed receipts could not be permanently deleted.",
            "error": str(error),
        }), 500

    for file_path in file_paths:
        try:
            os.remove(file_path)
        except OSError:
            app.logger.warning("Could not remove receipt file: %s", file_path)

    return jsonify({
        "success": True,
        "deleted_count": deleted_count,
        "message": "All trashed receipts were permanently deleted.",
    }), 200

@app.route('/accounts/receipts/clear-active', methods=['POST'])
def clear_active_receipts():

    # ==========================================================
    # CHECK WHO IS PERFORMING THE ACTION
    # ==========================================================

    department = str(

        session.get('department_id')
        or session.get('department')
        or session.get('role')
        or ''

    ).strip().lower()


    # ==========================================================
    # AUTHORIZE ACCOUNTS AND SYSTEMS ONLY
    # ==========================================================

    if department not in {

        'accounts',
        'system',
        'systems'

    }:

        return jsonify({

            'success': False,

            'message': (
                'You are not allowed to perform this action. '
                'Only Accounts and Systems departments can clear active receipts.'
            )

        }), 403


    try:

        # ======================================================
        # GET ONLY ACTIVE RECEIPTS
        # ======================================================

        active_receipts = Receipt.query.filter(

            db.or_(

                Receipt.is_deleted == False,

                Receipt.is_deleted.is_(None)

            )

        ).all()


        # ======================================================
        # CHECK IF THERE IS ANYTHING TO CLEAR
        # ======================================================

        if not active_receipts:

            return jsonify({

                'success': True,

                'cleared_count': 0,

                'message': 'There are no active receipts to clear.'

            })


        # ======================================================
        # GET CURRENT USER/OFFICER
        # ======================================================

        current_user = str(

            session.get('officer_name')
            or session.get('username')
            or session.get('user_id')
            or session.get('department_id')
            or 'Accounts'

        ).strip()


        # ======================================================
        # MOVE ALL ACTIVE RECEIPTS TO TRASH
        # ======================================================

        for receipt in active_receipts:

            receipt.is_deleted = True

            receipt.deleted_at = datetime.utcnow()

            receipt.deleted_by = current_user


        # ======================================================
        # SAVE ALL CHANGES
        # ======================================================

        db.session.commit()


        # ======================================================
        # SUCCESS RESPONSE
        # ======================================================

        return jsonify({

            'success': True,

            'cleared_count': len(active_receipts),

            'message': (
                f'{len(active_receipts)} receipt(s) successfully cleared '
                'from the active table.'
            )

        })


    except Exception as error:

        db.session.rollback()

        app.logger.exception(
            'Could not clear all active receipts'
        )


        return jsonify({

            'success': False,

            'message': (
                f'Error clearing active receipts: {str(error)}'
            )

        }), 500

def accounts_or_system_can_manage_receipts():
    department = str(
        session.get("department_id")
        or session.get("department")
        or session.get("role")
        or ""
    ).strip().lower()

    return department in {"accounts", "system", "systems"}


def deny_receipt_action():
    return jsonify({
        "success": False,
        "message": RECEIPT_ACTION_DENIED_MESSAGE,
    }), 403


def format_signed_time(value):
    if not value:
        return "Pending"

    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y %H:%M:%S")

    return str(value)

from datetime import datetime
from zoneinfo import ZoneInfo

@app.route(
    "/sign_student/<student_id>/<department>",
    methods=["GET", "POST"]
)
def sign_student(student_id, department):

    # ==========================================================
    # NORMALIZE DEPARTMENT
    # ==========================================================

    department = str(department).strip().lower()


    # ==========================================================
    # CHECK LOGGED-IN DEPARTMENT
    # ==========================================================

    logged_department = str(
        session.get("department_id", "")
    ).strip().lower()


    if logged_department != department:

        return "Unauthorized", 403


    # ==========================================================
    # FIND STUDENT
    # ==========================================================

    student = db.session.get(
        Student,
        student_id
    )


    if not student:

        flash(
            "Student not found.",
            "error"
        )

        return redirect(
            url_for("department_dashboard")
        )


    # ==========================================================
    # GET OFFICER NAME
    # ==========================================================

    officer_name = str(
        session.get("officer_name", "")
    ).strip()


    if not officer_name:

        flash(
            "Your officer name is missing. Please log in again.",
            "error"
        )

        return redirect(
            url_for("department_login")
        )


    # ==========================================================
    # PROCESS SIGNATURE
    # ==========================================================

    if request.method == "POST":


        # ------------------------------------------------------
        # IMPORTANT:
        # SAVE A REAL PYTHON DATETIME OBJECT TO DATABASE
        # ------------------------------------------------------

        signing_time = datetime.now(
            ZoneInfo("Africa/Lusaka")
        ).replace(
            tzinfo=None
        )


        # ------------------------------------------------------
        # GET DRAWN SIGNATURE AND APPROVAL
        # ------------------------------------------------------

        signature = request.form.get(
            "signature"
        )

        approval = request.form.get(
            "approval"
        )


        # ------------------------------------------------------
        # CHECK SIGNATURE
        # ------------------------------------------------------

        if not signature:

            flash(
                "Please draw your signature.",
                "error"
            )

            return redirect(
                request.url
            )


        # ------------------------------------------------------
        # CHECK APPROVAL
        # ------------------------------------------------------

        if not approval:

            flash(
                "Please draw your approval mark.",
                "error"
            )

            return redirect(
                request.url
            )


        # ------------------------------------------------------
        # SAVE SIGNATURE IMAGE
        # ------------------------------------------------------

        signature_path = save_signature_image(
            signature,
            department
        )


        # ------------------------------------------------------
        # SAVE APPROVAL IMAGE
        # ------------------------------------------------------

        approval_path = save_approval_image(
            approval,
            department
        )


        # ======================================================
        # SAVE DEPARTMENT INFORMATION
        # ======================================================

        if department == "reception":

            student.reception_signature = signature_path

            student.reception_clear = approval_path

            student.signed_reception = True

            student.reception_officer_name = officer_name

            student.reception_signed_time = signing_time


        elif department == "library":

            student.library_signature = signature_path

            student.library_clear = approval_path

            student.signed_library = True

            student.library_officer_name = officer_name

            student.library_signed_time = signing_time


        elif department == "admission":

            student.admissions_signature = signature_path

            student.admissions_clear = approval_path

            student.signed_admission = True

            student.admissions_officer_name = officer_name

            student.admissions_signed_time = signing_time


        elif department == "accounts":

            student.accounts_signature = signature_path

            student.accounts_clear = approval_path

            student.signed_accounts = True

            student.accounts_officer_name = officer_name

            student.accounts_signed_time = signing_time


        elif department == "systems":

            student.systems_signature = signature_path

            student.systems_clear = approval_path

            student.signed_systems = True

            student.systems_officer_name = officer_name

            student.systems_signed_time = signing_time


        elif department == "adosa":

            student.adosa_signature = signature_path

            student.adosa_clear = approval_path

            student.signed_adosa = True

            student.adosa_officer_name = officer_name

            student.adosa_signed_time = signing_time


        else:

            flash(
                "Invalid department.",
                "error"
            )

            return redirect(
                url_for("department_dashboard")
            )


        # ======================================================
        # FORMAT TIME ONLY FOR DISPLAY
        # ======================================================

        display_time = signing_time.strftime(
            "%d-%m-%Y %H:%M:%S"
        )


        # ======================================================
        # CREATE STUDENT NOTIFICATION
        # ======================================================

        department_name = department.upper()


        notification = SystemNotification(

            recipient_type="student",

            recipient_id=str(
                student.id
            ),

            student_email=student.email,

            title=(
                f"{department_name} Clearance Approved"
            ),

            message=(
                f"{department_name} Department signed "
                f"your clearance. Officer: "
                f"{officer_name}. Time: "
                f"{display_time}."
            ),

            is_read=False,
        )


        db.session.add(
            notification
        )


        # ======================================================
        # SAVE EVERYTHING TO DATABASE
        # ======================================================

        try:

            db.session.commit()


        except Exception as database_error:

            db.session.rollback()

            app.logger.exception(
                "Signature database save failed"
            )

            print(
                f"Signature database save failed: "
                f"{database_error}"
            )

            flash(
                "The signature could not be saved. "
                "Please try again.",
                "error"
            )

            return redirect(
                request.url
            )


        # ======================================================
        # COUNT SIGNED DEPARTMENTS
        # ======================================================

        signed_count = sum([

            bool(
                student.signed_accounts
            ),

            bool(
                student.signed_reception
            ),

            bool(
                student.signed_library
            ),

            bool(
                student.signed_admission
            ),

            bool(
                student.signed_systems
            ),

            bool(
                student.signed_adosa
            )

        ])


        clearance_complete = (
            signed_count == 6
        )


        # ======================================================
        # SEND EMAIL NOTIFICATION
        # ======================================================

        try:

            email_sent = send_department_signature_email(

                student=student,

                department=department,

                signed_count=signed_count,

                total_departments=6,

                clearance_complete=clearance_complete,

            )


            if email_sent:

                print(
                    f"{department_name} signature email "
                    f"sent to {student.email}"
                )


            else:

                print(
                    f"{department_name} signature email "
                    f"was not sent to {student.email}"
                )


        except Exception as email_error:

            print(
                f"{department_name} signature email "
                f"failed for {student.email}: "
                f"{email_error}"
            )


        # ======================================================
        # SUCCESS MESSAGE
        # ======================================================

        flash(

            f"Electronic Signature and Approval saved "
            f"successfully by {officer_name}.",

            "success"

        )


        # ======================================================
        # RETURN TO DEPARTMENT DASHBOARD
        # ======================================================

        return redirect(
            url_for("department_dashboard")
        )


    # ==========================================================
    # DISPLAY E-SIGNATURE PAGE
    # ==========================================================

    return render_template(

        "e_signature.html",

        student=student,

        department=department.upper(),

        officer_name=officer_name,

    )

@app.route("/check_notifications", methods=["GET"])
def check_notifications():
    """Return the current student's departmental signing updates."""

    student_id = session.get("student_id")

    if not student_id:
        return jsonify([])

    student = db.session.get(Student, student_id)

    if not student:
        return jsonify([])

    departments = [
        ("Reception", "signed_reception"),
        ("Library", "signed_library"),
        ("Admission", "signed_admission"),
        ("Accounts", "signed_accounts"),
        ("Systems", "signed_systems"),
        ("Assistant Dean Student Affairs", "signed_adosa"),
    ]

    notifications = []

    for department_name, signed_field in departments:
        is_signed = bool(getattr(student, signed_field, False))

        if is_signed:
            notifications.append({
                "title": f"{department_name} Department",
                "message": (
                    f"{department_name} has signed your clearance form."
                ),
                "status": "approved",
                "department": department_name,
            })
        else:
            notifications.append({
                "title": f"{department_name} Department",
                "message": (
                    f"Your clearance is still waiting for {department_name}."
                ),
                "status": "pending",
                "department": department_name,
            })

    return jsonify(notifications)

@app.route('/health-payment-history/<student_id>')
def health_payment_history(student_id):
    student = Student.query.get_or_404(student_id)

    semester1_payments = (
        PaymentHistory.query
        .filter_by(student_id=student.id, term='Semester 1')
        .order_by(PaymentHistory.payment_date.desc())
        .all()
    )

    semester2_payments = (
        PaymentHistory.query
        .filter_by(student_id=student.id, term='Semester 2')
        .order_by(PaymentHistory.payment_date.desc())
        .all()
    )

    return render_template(
        'health_payment_history.html',
        student=student,
        semester1_payments=semester1_payments,
        semester2_payments=semester2_payments,
    )
@app.route('/forgot-password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'POST':
        student_id = (request.form.get('student_id') or '').strip()
        email = (request.form.get('email') or '').strip().lower()

        if not student_id or not email:
            flash('Please enter your Student ID and registered email address.')
            return render_template('forgot_password.html')

        clean_expired_reset_codes()

        student = Student.query.filter(
            Student.id == student_id,
            func.lower(Student.email) == email
        ).first()

        if student:
            # Invalidate older unused codes for this student.
            PasswordResetCode.query.filter_by(
                student_id=student.id,
                used_at=None
            ).update(
                {'used_at': datetime.utcnow()},
                synchronize_session=False
            )

            code = generate_reset_code()

            reset_record = PasswordResetCode(
                student_id=student.id,
                code_hash=hash_reset_code(code),
                expires_at=datetime.utcnow() + timedelta(minutes=10),
                attempts=0
            )

            db.session.add(reset_record)
            db.session.commit()

            message = Message(
                subject='Rockview University password reset code',
                recipients=[student.email]
            )
            message.body = (
                f'Hello {student.name},\n\n'
                f'Your Rockview University password reset code is: {code}\n\n'
                'This code expires in 10 minutes and can be used only once.\n\n'
                'Rockview University Student Services'
            )
            mail.send(message)

        # Keep this message general so account details are not exposed.
        flash(
            'If the Student ID and email match our records, '
            'a reset code has been sent to the registered email.'
        )
        return redirect(url_for('verify_reset_code', student_id=student_id))

    return render_template('forgot_password.html')


@app.route('/verify-reset-code/<student_id>', methods=['GET', 'POST'])
def verify_reset_code(student_id):
    student = Student.query.get(student_id)

    if not student:
        flash('The reset request is invalid or has expired.')
        return redirect(url_for('forgot_password'))

    reset_record = PasswordResetCode.query.filter_by(
        student_id=student.id,
        used_at=None
    ).order_by(
        PasswordResetCode.created_at.desc()
    ).first()

    if not reset_record or reset_record.expires_at < datetime.utcnow():
        flash('The reset code has expired. Please request a new code.')
        return redirect(url_for('forgot_password'))

    if reset_record.attempts >= 5:
        flash('Too many incorrect attempts. Please request a new code.')
        return redirect(url_for('forgot_password'))

    if request.method == 'POST':
        code = ''.join((request.form.get('code') or '').split())
        reset_record.attempts += 1
        db.session.commit()

        if (
            len(code) != 6
            or not code.isdigit()
            or hash_reset_code(code) != reset_record.code_hash
        ):
            remaining = max(0, 5 - reset_record.attempts)
            flash(f'Incorrect reset code. Attempts remaining: {remaining}.')
            return render_template(
                'verify_reset_code.html',
                student_id=student.id
            )

        session['password_reset_student_id'] = student.id
        session['password_reset_record_id'] = reset_record.id
        return redirect(url_for('reset_password'))

    return render_template(
        'verify_reset_code.html',
        student_id=student.id
    )


@app.route('/reset-password', methods=['GET', 'POST'])
def reset_password():
    student_id = session.get('password_reset_student_id')
    reset_record_id = session.get('password_reset_record_id')

    if not student_id or not reset_record_id:
        flash('Please request a new password reset code.')
        return redirect(url_for('forgot_password'))

    student = Student.query.get(student_id)
    reset_record = PasswordResetCode.query.get(reset_record_id)

    if not student or not reset_record:
        session.pop('password_reset_student_id', None)
        session.pop('password_reset_record_id', None)
        flash('The reset request is invalid.')
        return redirect(url_for('forgot_password'))

    if (
        reset_record.used_at is not None
        or reset_record.expires_at < datetime.utcnow()
    ):
        session.pop('password_reset_student_id', None)
        session.pop('password_reset_record_id', None)
        flash('The reset code has expired. Please request a new code.')
        return redirect(url_for('forgot_password'))

    if request.method == 'POST':
        new_password = request.form.get('password') or ''
        confirm_password = request.form.get('confirm_password') or ''

        if len(new_password) < 8:
            flash('Your new password must contain at least 8 characters.')
            return render_template('reset_password.html')

        if new_password != confirm_password:
            flash('The passwords do not match.')
            return render_template('reset_password.html')

        # Update only the existing account password.
        student.set_password(new_password)
        reset_record.used_at = datetime.utcnow()
        db.session.commit()

        # Open the student's original dashboard account.
        session.pop('password_reset_student_id', None)
        session.pop('password_reset_record_id', None)
        session['student_id'] = student.id
        session['student_full_name'] = student.name

        flash('Your password has been changed successfully.')
        return redirect(url_for('dashboard'))

    return render_template('reset_password.html')

   # ==========================================================
# APPLICATION STARTUP
# ==========================================================

# ==========================================================
# APPLICATION STARTUP
# ==========================================================

if __name__ == "__main__":
    with app.app_context():
        db.create_all()

    port = int(os.environ.get("PORT", 5000))

    socketio.run(
        app,
        host="0.0.0.0",
        port=port,
        debug=False,
    )
